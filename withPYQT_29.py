"""
Decline Curve Analysis Tool

This application provides hierarchical well selection using a tree structure with checkboxes.
Users can select single or multiple wells, or aggregate data at formation, field, or district level.
Production data is summed for aggregated selections.

Key Features:
- Tree-based hierarchical selection (District → Field → Formation → Well)
- Automatic aggregation via sum for multiple selections
- Analysis on single or aggregated well data
"""

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QComboBox, QPushButton, QCheckBox, QLineEdit, QTextEdit,
                             QTabWidget, QGroupBox, QGridLayout, QRadioButton, QButtonGroup,
                             QMessageBox, QSplitter, QFrame, QListWidget, QListWidgetItem, QDialog, QDialogButtonBox,
                             QSizePolicy, QTreeWidget, QTreeWidgetItem, QStyleFactory, QFileDialog,
                             QInputDialog, QMenuBar, QMenu)
from PyQt6.QtCore import Qt, pyqtSignal, QEvent, QTimer
from PyQt6.QtGui import QFont, QPalette, QColor, QCursor, QIcon, QAction
import qtawesome as qta
import sys
import pandas as pd
import numpy as np
from scipy.optimize import curve_fit
from scipy.signal import savgol_filter, find_peaks
from scipy.stats import linregress
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg, NavigationToolbar2QT
import calendar
from bayes_opt import BayesianOptimization
import json
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Default parameter ranges for optimization
DEFAULT_FILTER_RANGES = {
    'threshold': (1.5, 2.2)
}

# Fixed constant for outlier iterations
MAX_OUTLIER_ITERATIONS = 10

DEFAULT_AUTO_START_RANGES = {
    'smooth_window': (5, 25),
    'min_segment_length': (6, 18),
    'change_sensitivity': (0.1, 0.5),
    'stability_window': (3, 12),
    'min_decline_rate': (0.05, 0.3)
}

int_keys = {
    'smooth_window', 'min_segment_length', 'stability_window'
}

# --- Arps model ---
def arps_hyperbolic(t, qi, Di, b):
    """Arps Hyperbolic model equation"""
    # Handle exponential decline case (b = 0)
    if abs(b) < 1e-10:  # Treat very small b as exponential
        return qi * np.exp(-Di * t)
    # Standard hyperbolic decline
    return qi / (1 + b * Di * t) ** (1/b)

def fit_arps_hyperbolic(t, q, fixed_qi=None):
    """Fit Arps Hyperbolic model to production data"""
    if fixed_qi is not None:
        def model_fixed_qi(t, Di, b):
            return arps_hyperbolic(t, fixed_qi, Di, b)
        p0 = [0.1, 0.5]
        try:
            popt_2, pcov_2 = curve_fit(
                model_fixed_qi, t, q, p0=p0,
                bounds=([1e-6, 0], [2, 2])
            )
            popt = [fixed_qi, popt_2[0], popt_2[1]]
            pcov = np.zeros((3, 3))
            pcov[1:, 1:] = pcov_2
            return popt, pcov
        except Exception as e:
            print(f"Model fitting with fixed qi failed: {e}")
            return None, None
    else:
        p0 = [q[0], 0.1, 0.5]
        try:
            popt, pcov = curve_fit(
                arps_hyperbolic, t, q, p0=p0,
                bounds=([1e-6, 1e-6, 0], [np.inf, 2, 2])
            )
            return popt, pcov
        except Exception as e:
            print(f"Model fitting failed: {e}")
            return None, None

def filter_outliers_iterative(t, q, threshold=3, max_iterations=MAX_OUTLIER_ITERATIONS, fixed_qi=None):
    mask = np.ones(len(t), dtype=bool)
    n = len(t)
    if n < 3:
        print(f"Data has only {n} points, skipping outlier removal to ensure at least 3 points.")
        popt, pcov = fit_arps_hyperbolic(t, q, fixed_qi=fixed_qi)
        return mask, popt
    for iteration in range(max_iterations):
        t_fit = t[mask]
        q_fit = q[mask]
        popt, pcov = fit_arps_hyperbolic(t_fit, q_fit, fixed_qi=fixed_qi)
        if popt is None:
            print(f"Fitting failed at iteration {iteration}")
            break
        q_pred = arps_hyperbolic(t, *popt)
        residuals = q - q_pred
        residuals_fit = residuals[mask]
        std_residuals = np.std(residuals_fit)
        new_mask = np.abs(residuals) <= threshold * std_residuals
        if np.sum(new_mask) < 3:
            print(f"Iteration {iteration + 1}: Would result in {np.sum(new_mask)} points, skipping outlier removal.")
            popt, pcov = fit_arps_hyperbolic(t[mask], q[mask], fixed_qi=fixed_qi)
            return mask, popt
        last_non_outlier_idx = -1
        for i in range(n - 1, -1, -1):
            if new_mask[i]:
                last_non_outlier_idx = i
                break
        if last_non_outlier_idx >= 0 and last_non_outlier_idx < n - 1:
            has_trailing_outliers = not new_mask[n - 1]
            if has_trailing_outliers:
                for i in range(last_non_outlier_idx + 1, n):
                    new_mask[i] = True
                print(f"Iteration {iteration + 1}: Protected {n - last_non_outlier_idx - 1} trailing data points from outlier removal")
        if np.array_equal(mask, new_mask):
            print(f"Converged after {iteration + 1} iterations")
            break
        n_removed = np.sum(mask) - np.sum(new_mask)
        print(f"Iteration {iteration + 1}: Removed {n_removed} outliers, std = {std_residuals:.2f}")
        mask = new_mask
    popt, pcov = fit_arps_hyperbolic(t[mask], q[mask], fixed_qi=fixed_qi)
    return mask, popt

def create_arps_plot(df_well, t, q, mask, popt, well_id, df_full, q_original, q_smoothed, 
                     outlier_threshold, forecast_avg_points=1, start_method="", 
                     show_outliers=True, show_pre_decline=True, show_forecast=True, 
                     show_inliers=True, show_channel=False, forecast_duration=60,
                     reference_models=None):
    if popt is None:
        return None
    fig, ax = plt.subplots(figsize=(12, 7))
    q_pred = arps_hyperbolic(t, *popt)
    residuals = q - q_pred
    avg_std = np.std(residuals)
    channel_width = avg_std 
    upper_bound = q_pred + channel_width
    lower_bound = q_pred - channel_width
    lower_bound = np.maximum(lower_bound, 0)
    last_date = df_well['Prod_Date'].max()
    forecast_months = forecast_duration
    forecast_dates = pd.date_range(start=last_date + pd.DateOffset(months=1),
                                   periods=forecast_months, freq='MS')
    last_t = t[-1]
    # Start forecast from the next time step to ensure continuity
    forecast_t = last_t + np.arange(1, forecast_months + 1)
    qi_original, Di, b = popt
    forecast_q = arps_hyperbolic(forecast_t, qi_original, Di, b)
    
    # Ensure continuity by checking the rate at the transition point
    last_historical_rate = q[mask][-1]
    first_forecast_rate = forecast_q[0]
    print(f"\nContinuity check:")
    print(f"Last historical rate: {last_historical_rate:.2f} bbl/day")
    print(f"First forecast rate (before adjustment): {first_forecast_rate:.2f} bbl/day")
    # Calculate the fitted model rate at the last historical point
    fitted_rate_at_last_t = arps_hyperbolic(last_t, qi_original, Di, b)
    
    if forecast_avg_points > 1:
        n_points = min(forecast_avg_points, np.sum(mask))
        initial_rate = np.mean(q[mask][-n_points:])
        print(f"\nUsing average of last {n_points} rates: {initial_rate:.2f} bbl/day")
        print(f"Fitted model rate at last point: {fitted_rate_at_last_t:.2f} bbl/day")
        rate_difference = initial_rate - fitted_rate_at_last_t
        forecast_q += rate_difference
        forecast_q = np.maximum(forecast_q, 0)
        print(f"Applied vertical shift of {rate_difference:.2f} bbl/day to match initial rate")
    elif forecast_avg_points == 0:
        # Use pure model forecast - ensure continuity by starting from fitted rate at last_t
        print(f"\nUsing fitted model forecast (no adjustment)")
        print(f"Fitted model rate at last point: {fitted_rate_at_last_t:.2f} bbl/day")
        # No adjustment needed - forecast already starts from the correct point
    else:
        initial_rate = q[mask][-1]
        print(f"\nUsing last historical rate: {initial_rate:.2f} bbl/day")
        print(f"Fitted model rate at last point: {fitted_rate_at_last_t:.2f} bbl/day")
        rate_difference = initial_rate - fitted_rate_at_last_t
        forecast_q += rate_difference
        forecast_q = np.maximum(forecast_q, 0)
        print(f"Applied vertical shift of {rate_difference:.2f} bbl/day to match last historical rate")
    
    # Final continuity check
    print(f"Final continuity check:")
    print(f"Last historical rate: {last_historical_rate:.2f} bbl/day")
    print(f"First forecast rate (after adjustment): {forecast_q[0]:.2f} bbl/day")
    print(f"Rate difference at transition: {forecast_q[0] - last_historical_rate:.2f} bbl/day")
    print(f"Forecast parameters: qi={qi_original:.2f}, Di={Di:.4f}, b={b:.2f}")
    decline_start_idx = len(q_original) - len(q)
    
    # Create continuous date range for smooth model curve
    first_date = df_well['Prod_Date'].min()
    last_historical_date = df_well['Prod_Date'].max()
    all_dates = pd.date_range(start=first_date, end=last_historical_date, freq='MS')
    
    # Create continuous time array for model prediction
    t_continuous = np.array([(d - first_date).days // 30 for d in all_dates])
    q_pred_continuous = arps_hyperbolic(t_continuous, *popt)
    
    if show_pre_decline and decline_start_idx > 0:
        ax.plot(df_full['Prod_Date'].iloc[:decline_start_idx], 
                q_original[:decline_start_idx], 'o', color='red', 
                label='Pre-Decline Data', alpha=0.6, markersize=5)
    
    if show_inliers:
        ax.plot(df_well['Prod_Date'][mask], q[mask], 'o', color='blue', 
                label='Decline Phase Data (Inliers)', markersize=6)
    if show_outliers and np.sum(~mask) > 0:
        ax.plot(df_well['Prod_Date'][~mask], q[~mask], 'x', color='red', 
                label='Outliers (Excluded)', markersize=10, markeredgewidth=2)
    
    # Plot continuous model curve for all months
    ax.plot(all_dates, q_pred_continuous, '-', color='green', 
            label='Arps Hyperbolic Model', linewidth=2)
    
    if show_channel:
        # Create continuous channel bounds
        upper_bound_continuous = q_pred_continuous + channel_width
        lower_bound_continuous = np.maximum(q_pred_continuous - channel_width, 0)
        ax.fill_between(all_dates, lower_bound_continuous, upper_bound_continuous,
                        color='green', alpha=0.2, label=f'±{channel_width:.1f} bbl/day Channel')
    if show_forecast:
        forecast_years = forecast_duration / 12
        if forecast_years in list(range(1, 31)):
            forecast_label = f'{int(forecast_years)}-Year Forecast'
        else:
            forecast_label = f'{forecast_duration}-Month Forecast'
        ax.plot(forecast_dates, forecast_q, '--', color='blue',
                label=forecast_label, linewidth=2)
    
    # Plot reference models if provided
    if reference_models is not None and len(reference_models) > 0:
        colors = ['purple', 'orange', 'brown', 'pink', 'gray', 'olive', 'cyan', 'magenta']
        
        for i, reference_model in enumerate(reference_models):
            ref_Di = reference_model['Di']
            ref_b = reference_model['b']
            ref_name = reference_model.get('name', f'Reference {i+1}')
            color = colors[i % len(colors)]
            
            # For reference model, we should use the original qi from when the reference was saved
            # But since we don't store the original qi, we'll scale the reference model to match the current model's qi
            # This ensures the reference model is visible and comparable
            ref_qi = qi_original
            
            # Create time array from start of fitted model to end of forecast
            # Historical part: from t=0 to last_t
            ref_t_historical = t_continuous
            ref_q_historical = arps_hyperbolic(ref_t_historical, ref_qi, ref_Di, ref_b)
            
            # Plot the historical part of reference model
            ax.plot(all_dates, ref_q_historical, '-.', color=color,
                    label=f'Reference Model ({ref_name})', linewidth=2, alpha=0.8)
            
            # Only show forecast part if show_forecast is True
            if show_forecast:
                # Forecast part: from last_t+1 to last_t+forecast_months
                ref_t_forecast = last_t + np.arange(1, forecast_months + 1)
                ref_q_forecast = arps_hyperbolic(ref_t_forecast, ref_qi, ref_Di, ref_b)
                
                # Apply the same vertical shift as the main forecast (if any)
                fitted_rate_at_last_t_ref = arps_hyperbolic(last_t, ref_qi, ref_Di, ref_b)
                
                if forecast_avg_points > 1:
                    n_points = min(forecast_avg_points, np.sum(mask))
                    initial_rate = np.mean(q[mask][-n_points:])
                    rate_difference = initial_rate - fitted_rate_at_last_t_ref
                    ref_q_forecast += rate_difference
                    ref_q_forecast = np.maximum(ref_q_forecast, 0)
                elif forecast_avg_points == 0:
                    # Use pure model forecast - no adjustment needed
                    pass
                else:
                    initial_rate = q[mask][-1]
                    rate_difference = initial_rate - fitted_rate_at_last_t_ref
                    ref_q_forecast += rate_difference
                    ref_q_forecast = np.maximum(ref_q_forecast, 0)
                
                # Plot the forecast part of reference model
                ax.plot(forecast_dates, ref_q_forecast, '-.', color=color,
                        linewidth=2, alpha=0.8)
    
    ax.set_xlabel('Date')
    ax.set_ylabel('Oil Production Rate (bbl/day)')
    title = f'Well {well_id}'
    ax.set_title(title)
    ax.legend()
    ax.grid(True, alpha=0.3)
    ax.set_ylim(bottom=0)
    fig.tight_layout()
    return fig

def get_days_in_month(date):
    return calendar.monthrange(date.year, date.month)[1]

def find_optimal_start_date(
    df_well,
    rate_col='M_Oil_Prod',
    smooth_window=13,
    min_segment_length=12,
    change_sensitivity=0.3,
    stability_window=6,
    min_decline_rate=0.15
):
    """
    Optimized robust method to detect the optimal start date for decline curve analysis.
    Uses change point detection combined with statistical analysis of production trends.
    
    Parameters:
    -----------
    df_well : DataFrame
        Well production data
    rate_col : str
        Column name for production rate
    smooth_window : int
        Window size for smoothing (must be odd)
    min_segment_length : int
        Minimum length of production segment to consider
    change_sensitivity : float
        Sensitivity for detecting significant changes (0-1)
    stability_window : int
        Window to check for stable declining trend
    min_decline_rate : float
        Minimum decline rate to consider as true decline (fraction)
    
    Returns:
    --------
    int : Index of optimal start date
    """
    df = df_well.sort_values('Prod_Date').copy()
    df['Prod_Date'] = pd.to_datetime(df['Prod_Date'])
    df['days_in_month'] = df['Prod_Date'].apply(get_days_in_month)
    q = (df[rate_col] / df['days_in_month']).values.astype(float)
    n = len(q)
    
    if n == 0:
        return 0
    
    if n < min_segment_length:
        print(f"Warning: Only {n} months of data available, less than minimum segment length {min_segment_length}")
        return 0
    
    # Step 1: Smooth the data
    w = int(smooth_window)
    if w < 3:
        w = 3
    if w % 2 == 0:
        w += 1
    if w > n:
        w = n if n % 2 == 1 else n - 1
    
    try:
        q_smooth = savgol_filter(q, window_length=w, polyorder=2)
    except Exception:
        q_smooth = q.copy()
    
    print(f"\n=== Optimal Start Date Detection ===")
    print(f"Total data points: {n}")
    
    # OPTIMIZATION 1: Use strided/sampled evaluation for large datasets
    # For datasets > 60 points, evaluate every 3rd point initially, then refine
    if n > 60:
        stride = 3
        print(f"Using optimized sampling (stride={stride}) for large dataset")
    else:
        stride = 1
    
    # OPTIMIZATION 2: Pre-compute cumulative statistics for faster calculations
    # Cumulative sums for quick mean/std calculations
    cumsum = np.cumsum(q_smooth)
    cumsum_sq = np.cumsum(q_smooth ** 2)
    
    def fast_stats(start_idx, end_idx):
        """Fast mean and std calculation using cumulative sums"""
        n_points = end_idx - start_idx
        if n_points <= 0:
            return 0, 0, np.inf
        
        sum_vals = cumsum[end_idx-1] - (cumsum[start_idx-1] if start_idx > 0 else 0)
        sum_sq = cumsum_sq[end_idx-1] - (cumsum_sq[start_idx-1] if start_idx > 0 else 0)
        
        mean_val = sum_vals / n_points
        if mean_val == 0:
            return mean_val, 0, np.inf
        
        variance = (sum_sq / n_points) - (mean_val ** 2)
        std_val = np.sqrt(max(0, variance))
        cv = std_val / mean_val
        
        return mean_val, std_val, cv
    
    # OPTIMIZATION 3: Pre-compute regression coefficients for all segments
    # Using vectorized operations
    def fast_linear_regression(segment):
        """Fast linear regression using numpy"""
        n_seg = len(segment)
        if n_seg < 3:
            return 0, 0, 0, False
        
        t = np.arange(n_seg)
        t_mean = (n_seg - 1) / 2  # Mean of 0,1,2,...,n-1
        
        # Vectorized computation
        slope = np.sum((t - t_mean) * segment) / np.sum((t - t_mean) ** 2)
        intercept = np.mean(segment) - slope * t_mean
        
        # R-squared
        y_pred = slope * t + intercept
        ss_res = np.sum((segment - y_pred) ** 2)
        ss_tot = np.sum((segment - np.mean(segment)) ** 2)
        r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0
        
        is_declining = slope < 0
        decline_rate = abs(slope) / intercept if intercept > 0 else 0
        
        return slope, intercept, r_squared, is_declining, decline_rate
    
    # Step 2: Quick scan with stride
    candidates = []
    
    max_start = n - min_segment_length
    sample_indices = list(range(0, max_start + 1, stride))
    
    # Always include some key points if not already included
    if n > min_segment_length:
        # Include points after detected peaks
        peaks, _ = find_peaks(q_smooth, distance=6)
        for p in peaks:
            if p <= max_start and p not in sample_indices:
                sample_indices.append(p)
    
    # Sort indices
    sample_indices.sort()
    
    for start_idx in sample_indices:
        segment = q_smooth[start_idx:]
        segment_length = len(segment)
        
        # Fast statistics
        mean_rate, std_rate, cv = fast_stats(start_idx, n)
        
        # Fast regression
        slope, intercept, r_squared, is_declining, decline_rate = fast_linear_regression(segment)
        
        if not is_declining:
            continue  # Skip non-declining segments early
        
        # OPTIMIZATION 4: Simplified stability check
        # Only check at key intervals instead of all windows
        stability_score = 0
        if segment_length >= stability_window:
            # Sample 5-7 windows instead of all
            n_samples = min(7, segment_length - stability_window + 1)
            sample_step = max(1, (segment_length - stability_window) // n_samples)
            
            negative_count = 0
            total_count = 0
            
            for i in range(0, segment_length - stability_window + 1, sample_step):
                window = segment[i:i+stability_window]
                t_win = np.arange(stability_window)
                # Quick slope check without full regression
                win_slope = (window[-1] - window[0]) / (stability_window - 1)
                total_count += 1
                if win_slope < 0:
                    negative_count += 1
            
            stability_score = negative_count / total_count if total_count > 0 else 0
        
        # OPTIMIZATION 5: Simplified event detection
        event_score = 0
        if start_idx > 0:
            before_mean, _, _ = fast_stats(0, start_idx)
            after_mean = mean_rate
            
            # Quick checks only
            if before_mean > after_mean * (1 + change_sensitivity):
                event_score += 0.5
            
            if start_idx > 0 and q_smooth[start_idx-1] > after_mean * (1 + change_sensitivity):
                event_score += 0.3
        
        # Calculate composite score
        length_score = min(1.0, segment_length / (min_segment_length * 2))
        cv_score = max(0, 1 - min(cv, 1.0))
        decline_score = min(1.0, decline_rate / min_decline_rate)
        linearity_score = r_squared
        
        composite_score = (
            0.25 * decline_score +
            0.25 * stability_score +
            0.20 * linearity_score +
            0.15 * length_score +
            0.10 * cv_score +
            0.05 * event_score
        )
        
        candidates.append({
            'idx': start_idx,
            'date': df.iloc[start_idx]['Prod_Date'],
            'score': composite_score,
            'length': segment_length,
            'cv': cv,
            'decline_rate': decline_rate,
            'stability': stability_score,
            'r_squared': r_squared,
            'is_declining': is_declining
        })
    
    if not candidates:
        print("Warning: No declining segments found, using all data")
        return 0
    
    # Step 3: Refine around top candidate if using stride
    if stride > 1 and len(candidates) > 0:
        # Sort to find best
        candidates.sort(key=lambda x: x['score'], reverse=True)
        best_idx = candidates[0]['idx']
        
        # Check neighbors around best candidate
        for refine_idx in range(max(0, best_idx - stride), min(max_start + 1, best_idx + stride + 1)):
            if refine_idx == best_idx:
                continue
            
            segment = q_smooth[refine_idx:]
            segment_length = len(segment)
            mean_rate, std_rate, cv = fast_stats(refine_idx, n)
            slope, intercept, r_squared, is_declining, decline_rate = fast_linear_regression(segment)
            
            if not is_declining:
                continue
            
            # Quick scoring
            length_score = min(1.0, segment_length / (min_segment_length * 2))
            cv_score = max(0, 1 - min(cv, 1.0))
            decline_score = min(1.0, decline_rate / min_decline_rate)
            
            composite_score = (
                0.35 * decline_score +
                0.35 * r_squared +
                0.20 * length_score +
                0.10 * cv_score
            )
            
            candidates.append({
                'idx': refine_idx,
                'date': df.iloc[refine_idx]['Prod_Date'],
                'score': composite_score,
                'length': segment_length,
                'cv': cv,
                'decline_rate': decline_rate,
                'stability': 0,  # Skip for refinement
                'r_squared': r_squared,
                'is_declining': is_declining
            })
    
    # Final sort
    candidates.sort(key=lambda x: x['score'], reverse=True)
    
    # Display top candidates
    print(f"\nEvaluated {len(candidates)} candidate start dates")
    print(f"Top 5 candidates:")
    for i, cand in enumerate(candidates[:5]):
        print(f"  {i+1}. Date: {cand['date'].date()}, "
              f"Score: {cand['score']:.3f}, "
              f"Length: {cand['length']} months, "
              f"CV: {cand['cv']:.3f}, "
              f"Decline: {cand['decline_rate']:.4f}")
    
    # Select best candidate
    best = candidates[0]
    
    print(f"\n✓ Selected start date: {best['date'].date()}")
    print(f"  Index: {best['idx']}")
    print(f"  Score: {best['score']:.3f}")
    print(f"  Data points used: {best['length']}")
    print(f"  Decline rate: {best['decline_rate']:.4f}")
    print(f"  R²: {best['r_squared']:.3f}")
    
    return int(best['idx'])

def calculate_fit_quality(t, q, mask, popt):
    if popt is None:
        return 0
    q_pred = arps_hyperbolic(t[mask], *popt)
    q_actual = q[mask]
    ss_res = np.sum((q_actual - q_pred) ** 2)
    ss_tot = np.sum((q_actual - np.mean(q_actual)) ** 2)
    r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0
    mape = np.mean(np.abs((q_actual - q_pred) / q_actual)) * 100 if np.all(q_actual > 0) else 1e6
    mape_score = max(0, 1 - min(mape, 100) / 100)
    qi, Di, b = popt

    # === Penalize very low Di ===
    Di_penalty = 1.0
    if Di < 0.001:
        Di_penalty = 0.01
    elif Di < 0.01:
        Di_penalty = 0.3
    elif Di < 0.05:
        Di_penalty = 0.7

    decline_reasonableness = 1.0
    if Di > 2.0:
        decline_reasonableness = 0.1
    elif b > 1.5:
        decline_reasonableness = 0.7

    return (0.5 * r_squared + 0.3 * mape_score + 0.2 * decline_reasonableness) * Di_penalty

def calculate_trend_consistency(t, q, mask, popt):
    q_actual = q[mask]
    t_actual = t[mask]
    if len(q_actual) < 3:
        return 0
    if len(q_actual) >= 6:
        recent_q = q_actual[-6:]
        recent_t = t_actual[-6:]
        actual_slope = linregress(recent_t, recent_q).slope
    else:
        actual_slope = linregress(t_actual, q_actual).slope
    q_pred = arps_hyperbolic(t_actual, *popt)
    expected_slope = linregress(t_actual, q_pred).slope
    if np.isnan(actual_slope) or np.isnan(expected_slope):
        return 0
    slope_ratio = abs(actual_slope / expected_slope) if expected_slope != 0 else 1
    consistency = 1.0 / (1.0 + abs(slope_ratio - 1.0))
    return consistency

def compute_metrics(df_all, well_name, manual_start_idx=0, filter_params=None, auto_start_params=None, ignore_gaps=False, fixed_qi=None):
    df_well = df_all[df_all['Well_Name'] == well_name].copy()
    if df_well.empty:
        return {'score': -1e10, 'popt': None}
    df_well = df_well.sort_values('Prod_Date')
    start_idx = manual_start_idx
    df_well['Prod_Date'] = pd.to_datetime(df_well['Prod_Date'])
    df_well['days_in_month'] = df_well['Prod_Date'].apply(get_days_in_month)
    df_well['oil_prod_daily'] = df_well['M_Oil_Prod'] / df_well['days_in_month']
    df_well['t'] = (df_well['Prod_Date'] - df_well['Prod_Date'].min()).dt.days // 30
    if start_idx > 0:
        df_well = df_well.iloc[start_idx:].reset_index(drop=True)
    if not ignore_gaps:
        if len(df_well) > 1:
            date_diffs = df_well['Prod_Date'].diff()
            large_gaps = date_diffs > pd.Timedelta(days=365 * 10)
            if large_gaps.any():
                last_gap_idx = np.where(large_gaps.values)[0][-1]
                points_after_gap = len(df_well) - (last_gap_idx + 1)
                if points_after_gap >= 10:
                    gap_date_before = df_well.iloc[last_gap_idx]['Prod_Date']
                    gap_date_after = df_well.iloc[last_gap_idx + 1]['Prod_Date']
                    gap_years = (gap_date_after - gap_date_before).days / 365.25
                    print(f"\n{'!'*60}")
                    print(f"LARGE DATA GAP DETECTED:")
                    print(f"Gap of {gap_years:.1f} years between:")
                    print(f"  {gap_date_before.date()} and {gap_date_after.date()}")
                    print(f"Points after gap: {points_after_gap}")
                    print(f"Excluding first {last_gap_idx + 1} points before the gap")
                    print(f"{'!'*60}\n")
                    df_well = df_well.iloc[last_gap_idx + 1:].reset_index(drop=True)
    t = df_well['t'].values - df_well['t'].values[0]
    q = df_well['oil_prod_daily'].values
    mask, popt = filter_outliers_iterative(
        t, q, 
        threshold=filter_params.get('threshold', 2.0),
        fixed_qi=fixed_qi
    )
    if popt is None:
        return {'score': -1e10, 'popt': None}
    num_used = np.sum(mask)
    total_points = len(mask)
    mean_q = np.mean(q[mask])
    cv = np.std(q[mask]) / mean_q if mean_q > 0 else 1e6
    base_score = num_used / (1 + cv*4)
    data_utilization = num_used / total_points
    fit_quality = calculate_fit_quality(t, q, mask, popt)
    trend_consistency = calculate_trend_consistency(t, q, mask, popt)
    enhanced_score = (
        0.3 * base_score +
        0.1 * data_utilization +
        0.3 * fit_quality +
        0.3 * trend_consistency
    )
    return {
        'score': enhanced_score,
        'base_score': base_score,
        'cv': cv,
        'num_used': num_used,
        'total_points': total_points,
        'data_utilization': data_utilization,
        'fit_quality': fit_quality,
        'trend_consistency': trend_consistency,
        'popt': popt,
        'mask': mask
    }

def run_arps_for_well_auto(df_all, well_name, outlier_threshold=2, 
                           forecast_avg_points=6, manual_start_idx=None,
                           use_auto_detect=True, start_method="", 
                           auto_start_params=None, filter_params=None,
                           show_outliers=True, show_pre_decline=True, 
                           show_forecast=True, show_inliers=True, show_channel=False, 
                           forecast_duration=60, fixed_qi=None,
                           reference_models=None):
    df_well = df_all[df_all['Well_Name'] == well_name].copy()
    print(f"DEBUG: run_arps_for_well_auto called with well_name='{well_name}'")
    print(f"DEBUG: df_all has wells: {df_all['Well_Name'].unique()}")
    print(f"DEBUG: df_well empty: {df_well.empty}")
    if df_well.empty:
        return None, "No data found for well"
    df_well = df_well.sort_values('Prod_Date')
    if manual_start_idx is not None:
        start_idx = manual_start_idx
    elif use_auto_detect:
        start_idx = find_optimal_start_date(
            df_well,
            rate_col='M_Oil_Prod',
            smooth_window=auto_start_params.get('smooth_window', 13),
            min_segment_length=auto_start_params.get('min_segment_length', 12),
            change_sensitivity=auto_start_params.get('change_sensitivity', 0.3),
            stability_window=auto_start_params.get('stability_window', 6),
            min_decline_rate=auto_start_params.get('min_decline_rate', 0.15)
        )
    else:
        start_idx = 0
    df_well['Prod_Date'] = pd.to_datetime(df_well['Prod_Date'])
    df_well['days_in_month'] = df_well['Prod_Date'].apply(get_days_in_month)
    df_well['oil_prod_daily'] = df_well['M_Oil_Prod'] / df_well['days_in_month']
    df_well['t'] = (df_well['Prod_Date'] - df_well['Prod_Date'].min()).dt.days // 30
    df_full = df_well.copy()
    q_original_full = df_well['oil_prod_daily'].values.copy()
    n = len(q_original_full)
    w = min(7, n if n % 2 == 1 else n-1)
    w = max(3, w)
    try:
        q_smoothed_full = savgol_filter(q_original_full, window_length=w, polyorder=2)
    except Exception:
        q_smoothed_full = q_original_full.copy()
    if start_idx > 0:
        df_well = df_well.iloc[start_idx:].reset_index(drop=True)
        print(f"\n{'='*60}")
        print(f"Analysis starting from: {df_well.iloc[0]['Prod_Date'].date()}")
        print(f"Data points used: {len(df_well)}")
        print(f"{'='*60}\n")
    if len(df_well) > 1:
        date_diffs = df_well['Prod_Date'].diff()
        large_gaps = date_diffs > pd.Timedelta(days=365 * 10)
        print(f"\n--- Checking for large data gaps ---")
        print(f"Start method: {start_method if start_method else 'auto-detect'}")
        print(f"Auto detect enabled: {use_auto_detect}")
        print(f"Large gaps found: {large_gaps.any()}")
        print(f"--- End of gap check ---\n")
        if start_method not in ["Manual Start Date", "Manual Start Date and Initial Rate (Qi)"] and large_gaps.any():
            last_gap_idx = np.where(large_gaps.values)[0][-1]
            points_after_gap = len(df_well) - (last_gap_idx + 1)
            if points_after_gap >= 10:
                gap_date_before = df_well.iloc[last_gap_idx]['Prod_Date']
                gap_date_after = df_well.iloc[last_gap_idx + 1]['Prod_Date']
                gap_years = (gap_date_after - gap_date_before).days / 365.25
                print(f"\n{'!'*60}")
                print(f"LARGE DATA GAP DETECTED:")
                print(f"Gap of {gap_years:.1f} years between:")
                print(f"  {gap_date_before.date()} and {gap_date_after.date()}")
                print(f"Points after gap: {points_after_gap}")
                print(f"Excluding first {last_gap_idx + 1} points before the gap")
                print(f"{'!'*60}\n")
                df_well = df_well.iloc[last_gap_idx + 1:].reset_index(drop=True)
                start_idx += last_gap_idx + 1
        else:
            print(f"\nManual start selected — using all available points after {df_well.iloc[0]['Prod_Date'].date()} (ignoring gaps).")
    t = df_well['t'].values - df_well['t'].values[0]
    q = df_well['oil_prod_daily'].values
    print("--- Starting iterative outlier filtering ---")
    mask, popt = filter_outliers_iterative(
        t, q, 
        threshold=filter_params.get('threshold', outlier_threshold),
        fixed_qi=fixed_qi
    )
    if popt is not None:
        n_outliers = np.sum(~mask)
        print(f"\n=== Final Results for well {well_name} ===")
        print(f"Total outliers removed: {n_outliers} out of {len(mask)} points")
        print(f"Fitted parameters:")
        print(f"qi = {popt[0]:.2f}, Di = {popt[1]:.4f}, b = {popt[2]:.2f}")
        cv = np.std(q[mask]) / np.mean(q[mask])
        print(f"\nCoefficient of Variation: {cv:.3f}")
        if cv > 0.4:
            print("⚠ WARNING: High variability detected. Forecast may be unreliable.")
        elif cv > 0.3:
            print("⚠ CAUTION: Moderate variability. Review forecast carefully.")
        else:
            print("✓ Good data quality for decline curve analysis.")
    fig = create_arps_plot(df_well, t, q, mask, popt, well_name, df_full, 
                           q_original_full, q_smoothed_full, outlier_threshold, 
                           forecast_avg_points, start_method,
                           show_outliers, show_pre_decline, show_forecast, 
                           show_inliers, show_channel, forecast_duration,
                           reference_models=reference_models)
    
    results = ""
    if popt is not None:
        n_outliers = np.sum(~mask)
        results += f"\n=== Final Results for well {well_name} ===\n"
        results += f"Total outliers removed: {n_outliers} out of {len(mask)} points\n"
        results += f"Fitted parameters:\n"
        results += f"qi = {popt[0]:.2f}, Di = {popt[1]:.4f}, b = {popt[2]:.2f}\n"
        cv = np.std(q[mask]) / np.mean(q[mask])
        results += f"\nCoefficient of Variation: {cv:.3f}\n"
        if cv > 0.4:
            results += "⚠ WARNING: High variability detected. Forecast may be unreliable.\n"
        elif cv > 0.3:
            results += "⚠ CAUTION: Moderate variability. Review forecast carefully.\n"
        else:
            results += "✓ Good data quality for decline curve analysis.\n"
    
    return fig, results

class DeclineCurveApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Decline Curve Analysis Tool")
        self.setGeometry(100, 100, 1600, 1000)
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f8f9fa;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            
            /* Menu Bar Styling */
            QMenuBar {
                background-color: #ffffff;
                border-bottom: 1px solid #dee2e6;
                padding: 4px;
                font-size: 12px;
                font-weight: 500;
            }
            QMenuBar::item {
                background-color: transparent;
                padding: 8px 12px;
                border-radius: 4px;
                margin: 2px;
            }
            QMenuBar::item:selected {
                background-color: #e9ecef;
                color: #495057;
            }
            QMenuBar::item:pressed {
                background-color: #dee2e6;
            }
            QMenu {
                background-color: #ffffff;
                border: 1px solid #dee2e6;
                border-radius: 6px;
                padding: 4px;
                font-size: 12px;
            }
            QMenu::item {
                padding: 8px 16px;
                border-radius: 4px;
                margin: 1px;
            }
            QMenu::item:selected {
                background-color: #007bff;
                color: white;
            }
            QMenu::separator {
                height: 1px;
                background-color: #dee2e6;
                margin: 4px 8px;
            }
            
            /* Typography */
            QLabel {
                color: #495057;
                font-size: 12px;
                font-weight: 400;
            }
            
            /* Group Boxes */
            QGroupBox {
                font-weight: 600;
                font-size: 13px;
                color: #495057;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 15px;
                background-color: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px 0 8px;
                background-color: #ffffff;
            }
            
            /* Buttons - Primary (Blue) */
            QPushButton {
                background-color: #007bff;
                border: none;
                color: white;
                padding: 10px 16px;
                text-align: center;
                font-size: 12px;
                font-weight: 500;
                border-radius: 6px;
                min-height: 20px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #004085;
            }
            QPushButton:disabled {
                background-color: #6c757d;
                color: #ffffff;
            }
            
            /* Buttons - Secondary (Gray) */
            QPushButton[class="secondary"] {
                background-color: #6c757d;
                color: white;
            }
            QPushButton[class="secondary"]:hover {
                background-color: #545b62;
            }
            QPushButton[class="secondary"]:pressed {
                background-color: #3d4449;
            }
            
            /* Buttons - Success (Green) */
            QPushButton[class="success"] {
                background-color: #28a745;
                color: white;
            }
            QPushButton[class="success"]:hover {
                background-color: #1e7e34;
            }
            QPushButton[class="success"]:pressed {
                background-color: #155724;
            }
            
            /* Buttons - Danger (Red) */
            QPushButton[class="danger"] {
                background-color: #dc3545;
                color: white;
            }
            QPushButton[class="danger"]:hover {
                background-color: #c82333;
            }
            QPushButton[class="danger"]:pressed {
                background-color: #bd2130;
            }
            
            /* Buttons - Warning (Orange) */
            QPushButton[class="warning"] {
                background-color: #ffc107;
                color: #212529;
            }
            QPushButton[class="warning"]:hover {
                background-color: #e0a800;
            }
            QPushButton[class="warning"]:pressed {
                background-color: #d39e00;
            }
            
            /* Primary Action Button (Run Analysis) */
            QPushButton[class="primary"] {
                background-color: #007bff;
                font-size: 12px;
                font-weight: 500;
                padding: 10px 16px;
                min-height: 20px;
            }
            QPushButton[class="primary"]:hover {
                background-color: #0056b3;
            }
            QPushButton[class="primary"]:pressed {
                background-color: #004085;
            }
            
            /* Tab Widget */
            QTabWidget::pane {
                border: 1px solid #dee2e6;
                background-color: #ffffff;
                border-radius: 8px;
            }
            QTabBar::tab {
                background-color: #f8f9fa;
                color: #495057;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-weight: 500;
                font-size: 12px;
            }
            QTabBar::tab:selected {
                background-color: #007bff;
                color: white;
            }
            QTabBar::tab:hover:!selected {
                background-color: #e9ecef;
            }
            
            /* Tree Widget */
            QTreeWidget {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: white;
                padding: 4px;
            }
            QTreeWidget::item {
                padding: 4px;
                border-radius: 4px;
            }
            QTreeWidget::item:selected {
                background-color: #007bff;
                color: white;
            }
            QTreeWidget::item:hover {
                background-color: #e9ecef;
            }
            
            /* Input Fields */
            QLineEdit {
                border: 1px solid #ced4da;
                border-radius: 6px;
                padding: 8px 12px;
                background-color: white;
                font-size: 12px;
                color: #495057;
            }
            QLineEdit:focus {
                border-color: #007bff;
            }
            QLineEdit:disabled {
                background-color: #e9ecef;
                color: #6c757d;
            }
            
            /* Checkboxes */
            QCheckBox {
                spacing: 8px;
                font-size: 12px;
                color: #495057;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border-radius: 4px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #ced4da;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #007bff;
                background-color: #007bff;
            }
            QCheckBox::indicator:hover {
                border-color: #007bff;
            }
            
            /* Radio Buttons */
            QRadioButton {
                spacing: 8px;
                font-size: 12px;
                color: #495057;
            }
            QRadioButton::indicator {
                width: 18px;
                height: 18px;
                border-radius: 9px;
            }
            QRadioButton::indicator:unchecked {
                border: 2px solid #ced4da;
                background-color: white;
            }
            QRadioButton::indicator:checked {
                border: 2px solid #007bff;
                background-color: #007bff;
            }
            QRadioButton::indicator:hover {
                border-color: #007bff;
            }
            
            /* Text Edit */
            QTextEdit {
                border: 1px solid #ced4da;
                border-radius: 6px;
                background-color: white;
                padding: 8px;
                font-size: 12px;
                color: #495057;
            }
            QTextEdit:focus {
                border-color: #007bff;
            }
            
            /* Combo Box */
            QComboBox {
                border: 1px solid #ced4da;
                border-radius: 6px;
                padding: 8px 12px;
                background-color: white;
                font-size: 12px;
                color: #495057;
                min-width: 100px;
            }
            QComboBox:focus {
                border-color: #007bff;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTIiIGhlaWdodD0iMTIiIHZpZXdCb3g9IjAgMCAxMiAxMiIgZmlsbD0ibm9uZSIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj4KPHBhdGggZD0iTTMgNEw2IDdMOSA0IiBzdHJva2U9IiM2Yzc1N2QiIHN0cm9rZS13aWR0aD0iMiIgc3Ryb2tlLWxpbmVjYXA9InJvdW5kIiBzdHJva2UtbGluZWpvaW49InJvdW5kIi8+Cjwvc3ZnPgo=);
            }
            
            /* Splitter */
            QSplitter::handle {
                background-color: #dee2e6;
                width: 3px;
                border-radius: 1px;
            }
            QSplitter::handle:hover {
                background-color: #007bff;
            }
        """)
        self.filter_ranges = DEFAULT_FILTER_RANGES.copy()
        self.auto_start_ranges = DEFAULT_AUTO_START_RANGES.copy()
        
        # Session storage for all well analyses
        self.session_analyses = {}  # Key: well_name, Value: analysis data dict
        self.saved_analyses = {}  # Track which analyses have been saved (well_name -> True/False)
        
        try:
            self.df_all = pd.read_csv('OFM202409.csv', low_memory=False)
            self.districts = sorted([d for d in self.df_all['District'].dropna().unique() if str(d).strip()])
            self.fields_by_district = {}
            self.formations_by_field = {}
            self.wells_by_formation = {}
            for district in self.districts:
                district_data = self.df_all[self.df_all['District'] == district]
                fields = sorted([f for f in district_data['Field'].dropna().unique() if str(f).strip()])
                self.fields_by_district[district] = fields
                for field in fields:
                    field_data = district_data[district_data['Field'] == field]
                    formations = sorted([fm for fm in field_data['Alias_Formation'].dropna().unique() if str(fm).strip()])
                    self.formations_by_field[f"{district}|{field}"] = formations
                    for formation in formations:
                        formation_data = field_data[field_data['Alias_Formation'] == formation]
                        wells = sorted([w for w in formation_data['Well_Name'].dropna().unique() if str(w).strip()])
                        self.wells_by_formation[f"{district}|{field}|{formation}"] = wells
            self.initializing = True  # Flag to prevent reset during initialization
            # Keep a reference to the full dataset for title computations
            self.df_master = self.df_all.copy()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data: {str(e)}")
            sys.exit(1)
            return

        # Create menu bar
        self.create_menu_bar()
        
        # Create main container
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(6)
        
        # Create main splitter for left controls and right plot area
        splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(splitter)
        
        # Left panel with tabbed controls
        left_widget = QWidget()
        left_widget.setMinimumWidth(500)
        left_widget.setMaximumWidth(700)
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 10, 0)
        
        # Create main tab widget for left panel
        self.main_tabs = QTabWidget()
        left_layout.addWidget(self.main_tabs)
        
        # Setup Tab (Well Selection + Start Date)
        self.create_setup_tab()
        
        # Parameters Tab (Optimize Parameters + Run/Reset buttons)
        self.create_parameters_tab()
        
        # Exports Tab
        self.create_exports_tab()
        
        # Chart frame (expanded to fill right side)
        self.canvas_frame = QWidget()
        self.canvas_layout = QVBoxLayout(self.canvas_frame)
        self.canvas_layout.setContentsMargins(0, 0, 0, 0)
        self.canvas_layout.setSpacing(0)
        
        splitter.addWidget(left_widget)
        splitter.addWidget(self.canvas_frame)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        
        # Prevent complete collapse of panes
        splitter.setCollapsible(0, False)  # Left panel cannot be collapsed
        splitter.setCollapsible(1, False)  # Right panel cannot be collapsed
        
        # Set minimum sizes to ensure usability
        left_widget.setMinimumWidth(500)  # Minimum width for left panel
        self.canvas_frame.setMinimumWidth(300)  # Minimum width for plot area

        # Store the currently applied selection
        self.applied_wells = []
        # Store aggregated data for chart click selection when multiple wells are selected
        self.df_aggregated_cache = None

        self.figure = None
        self.canvas = None
        self.toolbar = None
        self.current_analysis_params = None
        self.current_well_name = None
        self.click_selection_enabled = False
        self._loading_analysis = False  # Flag to prevent signal handlers during loading
        self.qi_selection_enabled = False
        self.click_cid = None
        self.qi_click_cid = None
        self.key_cid = None
        self.fixed_qi = None
        self.original_cursor = None
        self.event_filter_installed = False
        self.vertical_line = None  # Store reference to vertical line on chart
        
        # Initialization complete, enable well selection reset behavior
        self.initializing = False

    def create_menu_bar(self):
        """Create the menu bar with File and View menus"""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu('File')
        
        # Load Session action
        load_action = QAction(qta.icon('fa5s.folder-open'), 'Load Session...', self)
        load_action.setShortcut('Ctrl+O')
        load_action.setStatusTip('Load all analyses from a saved session file')
        load_action.triggered.connect(self.load_session)
        file_menu.addAction(load_action)
        
        # Save Session action
        save_action = QAction(qta.icon('fa5s.save'), 'Save Session...', self)
        save_action.setShortcut('Ctrl+S')
        save_action.setStatusTip('Save all analyses in the current session to a file')
        save_action.triggered.connect(self.save_session)
        file_menu.addAction(save_action)
        
        file_menu.addSeparator()
        
        # Exit action
        exit_action = QAction(qta.icon('fa5s.sign-out-alt'), 'Exit', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.setStatusTip('Exit the application')
        exit_action.triggered.connect(self.exit_application)
        file_menu.addAction(exit_action)
        
        # View menu
        view_menu = menubar.addMenu('View')
        
        # View Analyses action
        view_analyses_action = QAction(qta.icon('fa5s.list'), 'View Analyses...', self)
        view_analyses_action.setShortcut('Ctrl+V')
        view_analyses_action.setStatusTip('View and select from available analyses in the current session')
        view_analyses_action.triggered.connect(self.show_analyses_list)
        view_menu.addAction(view_analyses_action)

    def create_setup_tab(self):
        """Create the Setup tab with Well Selection and Start Date controls"""
        setup_tab = QWidget()
        setup_layout = QVBoxLayout(setup_tab)
        setup_layout.setSpacing(10)
        
        # Well Selection Group
        selection_group = QGroupBox("Well Selection")
        selection_layout = QVBoxLayout(selection_group)
        selection_layout.setSpacing(8)
        
        # Create tree widget with checkboxes
        self.well_tree = QTreeWidget()
        self.well_tree.setHeaderLabel("District → Field → Formation → Well")
        # Ensure checkbox indicators remain visible when the tree loses focus
        self.well_tree.setStyle(QStyleFactory.create("Fusion"))
        self.well_tree.setMinimumHeight(200)
        self.well_tree.setMaximumHeight(300)
        # Disable item selection/highlighting (only checkboxes are interactive)
        self.well_tree.setSelectionMode(QTreeWidget.SelectionMode.NoSelection)
        self.well_tree.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.well_tree.itemChanged.connect(self.on_tree_item_changed)
        
        # Populate the tree
        self.populate_well_tree()
        
        selection_layout.addWidget(self.well_tree)
        
        # Add buttons and status label
        buttons_layout = QHBoxLayout()
        
        self.apply_selection_btn = QPushButton("Apply")
        self.apply_selection_btn.clicked.connect(self.apply_selection)
        self.apply_selection_btn.setProperty("class", "success")
        buttons_layout.addWidget(self.apply_selection_btn)
        
        self.clear_selection_btn = QPushButton("Clear Selection")
        self.clear_selection_btn.clicked.connect(self.clear_selection)
        self.clear_selection_btn.setProperty("class", "secondary")
        buttons_layout.addWidget(self.clear_selection_btn)
        
        # Label to show current applied selection
        self.applied_selection_label = QLabel("No selection applied")
        self.applied_selection_label.setStyleSheet("""
            QLabel {
                color: #6c757d;
                font-style: italic;
                font-size: 11px;
            }
        """)
        buttons_layout.addWidget(self.applied_selection_label, 1)
        
        selection_layout.addLayout(buttons_layout)
        setup_layout.addWidget(selection_group)
        
        # Start Date Selection Method
        start_method_group = QGroupBox("Start Date Selection Method")
        start_method_layout = QVBoxLayout(start_method_group)
        start_method_layout.setSpacing(8)
        
        self.start_method_buttons = QButtonGroup()
        self.auto_select_radio = QRadioButton("Auto Select")
        self.auto_select_radio.setChecked(True)
        self.auto_select_radio.toggled.connect(lambda checked: self.update_vertical_line_from_inputs() if checked else None)
        self.start_method_buttons.addButton(self.auto_select_radio, 0)
        start_method_layout.addWidget(self.auto_select_radio)

        # Manual Start Date
        self.manual_frame = QGroupBox("Manual Start Date")
        manual_layout = QVBoxLayout(self.manual_frame)
        
        self.manual_start_radio = QRadioButton("Manual Start Date")
        self.manual_start_radio.toggled.connect(lambda checked: self.update_vertical_line_from_inputs() if checked else None)
        self.start_method_buttons.addButton(self.manual_start_radio, 1)
        manual_layout.addWidget(self.manual_start_radio)
        
        manual_date_layout = QHBoxLayout()
        manual_date_layout.addWidget(QLabel("Year:"))
        self.manual_year_edit = QLineEdit("2020")
        self.manual_year_edit.setMaximumWidth(60)
        self.manual_year_edit.textChanged.connect(self.update_vertical_line_from_inputs)
        self.manual_year_edit.editingFinished.connect(lambda: self.validate_manual_date())
        manual_date_layout.addWidget(self.manual_year_edit)
        manual_date_layout.addWidget(QLabel("Month:"))
        self.manual_month_edit = QLineEdit("1")
        self.manual_month_edit.setMaximumWidth(40)
        self.manual_month_edit.textChanged.connect(self.update_vertical_line_from_inputs)
        self.manual_month_edit.editingFinished.connect(lambda: self.validate_manual_date())
        manual_date_layout.addWidget(self.manual_month_edit)
        self.select_start_btn = QPushButton("Select on Chart")
        self.select_start_btn.clicked.connect(self.enable_chart_click_selection)
        self.select_start_btn.setProperty("class", "secondary")
        manual_date_layout.addWidget(self.select_start_btn)
        manual_date_layout.addStretch()
        manual_layout.addLayout(manual_date_layout)
        
        start_method_layout.addWidget(self.manual_frame)

        # Manual Start Date and Initial Rate (Qi)
        self.qi_frame = QGroupBox("Manual Start Date and Initial Rate (Qi)")
        qi_layout = QVBoxLayout(self.qi_frame)
        
        self.qi_radio = QRadioButton("Manual Start Date and Initial Rate (Qi)")
        self.qi_radio.toggled.connect(lambda checked: self.update_vertical_line_from_inputs() if checked else None)
        self.start_method_buttons.addButton(self.qi_radio, 2)
        qi_layout.addWidget(self.qi_radio)
        
        qi_date_layout = QHBoxLayout()
        qi_date_layout.addWidget(QLabel("Year:"))
        self.qi_year_edit = QLineEdit("2020")
        self.qi_year_edit.setMaximumWidth(60)
        self.qi_year_edit.textChanged.connect(self.update_vertical_line_from_inputs)
        self.qi_year_edit.editingFinished.connect(lambda: self.validate_qi_date())
        qi_date_layout.addWidget(self.qi_year_edit)
        qi_date_layout.addWidget(QLabel("Month:"))
        self.qi_month_edit = QLineEdit("1")
        self.qi_month_edit.setMaximumWidth(40)
        self.qi_month_edit.textChanged.connect(self.update_vertical_line_from_inputs)
        self.qi_month_edit.editingFinished.connect(lambda: self.validate_qi_date())
        qi_date_layout.addWidget(self.qi_month_edit)
        qi_date_layout.addStretch()
        qi_layout.addLayout(qi_date_layout)
        
        qi_rate_layout = QHBoxLayout()
        qi_rate_layout.addWidget(QLabel("Initial Rate (Qi) bbl/day:"))
        self.qi_value_edit = QLineEdit("")
        self.qi_value_edit.setMaximumWidth(100)
        qi_rate_layout.addWidget(self.qi_value_edit)
        self.select_qi_btn = QPushButton("Select on Chart")
        self.select_qi_btn.clicked.connect(self.enable_qi_click_selection)
        self.select_qi_btn.setProperty("class", "secondary")
        qi_rate_layout.addWidget(self.select_qi_btn)
        qi_rate_layout.addStretch()
        qi_layout.addLayout(qi_rate_layout)
        
        start_method_layout.addWidget(self.qi_frame)
        setup_layout.addWidget(start_method_group)
        
        # Primary action buttons in Setup tab
        button_layout = QHBoxLayout()
        self.run_btn = QPushButton("Run Analysis")
        self.run_btn.clicked.connect(self.run_analysis)
        self.run_btn.setProperty("class", "primary")
        button_layout.addWidget(self.run_btn)
        
        self.reset_btn = QPushButton("Reset")
        self.reset_btn.clicked.connect(self.reset_parameters)
        self.reset_btn.setProperty("class", "danger")
        button_layout.addWidget(self.reset_btn)
        setup_layout.addLayout(button_layout)
        
        # Add to main tabs
        self.main_tabs.addTab(setup_tab, "Analysis")

    def create_parameters_tab(self):
        """Create the Parameters tab with optimization controls and action buttons"""
        parameters_tab = QWidget()
        parameters_layout = QVBoxLayout(parameters_tab)
        parameters_layout.setSpacing(10)
        
        # Optimization controls
        opt_layout = QHBoxLayout()
        self.optimize_check = QCheckBox("Optimize Parameters")
        self.optimize_check.setChecked(True)
        opt_layout.addWidget(self.optimize_check)
        opt_config_btn = QPushButton("Configure Optimization Ranges")
        opt_config_btn.clicked.connect(self.open_range_config)
        opt_config_btn.setProperty("class", "secondary")
        opt_layout.addWidget(opt_config_btn)
        parameters_layout.addLayout(opt_layout)

        # Tabs for settings
        self.settings_tabs = QTabWidget()
        parameters_layout.addWidget(self.settings_tabs)
        self.create_general_tab()
        self.create_auto_start_tab()
        self.create_plot_options_tab()
        self.create_reference_tab()

        # Results text area
        self.results_text = QTextEdit()
        self.results_text.setMaximumHeight(150)
        self.results_text.setReadOnly(True)
        parameters_layout.addWidget(self.results_text)
        
        # Add to main tabs
        self.main_tabs.addTab(parameters_tab, "Setup")

    # =========================================================================
    # HELPER METHODS
    # =========================================================================
    
    def _cleanup_canvas_and_toolbar(self):
        """Helper method to clean up canvas and toolbar widgets"""
        if self.canvas:
            self.canvas_layout.removeWidget(self.canvas)
            self.canvas.deleteLater()
            self.canvas = None
        if self.toolbar:
            self.canvas_layout.removeWidget(self.toolbar)
            self.toolbar.deleteLater()
            self.toolbar = None
    
    def _display_figure(self, fig):
        """Helper method to display a matplotlib figure in the canvas"""
        if fig is None:
            return
        
        self._cleanup_canvas_and_toolbar()
        
        self.figure = fig
        self.canvas = FigureCanvasQTAgg(fig)
        self.toolbar = NavigationToolbar2QT(self.canvas, self.canvas_frame)
        self.canvas_layout.addWidget(self.toolbar)
        self.canvas_layout.addWidget(self.canvas)
        
        # Reset event filter flag for new canvas
        self.event_filter_installed = False
    
    def _get_well_data_for_selection(self):
        """Get the appropriate dataframe for chart selection (aggregated or single well)"""
        if len(self.applied_wells) > 1 and self.df_aggregated_cache is not None:
            return self.df_aggregated_cache.copy()
        else:
            well_name = self.get_primary_well()
            if not well_name:
                return None
            return self.df_all[self.df_all['Well_Name'] == well_name].copy()
    
    def _draw_vertical_line_at_date(self, selected_date):
        """Draw a vertical line at the specified date on the chart"""
        if self.figure is None:
            return
        
        # Remove any existing vertical line
        self.remove_vertical_line()
        
        # Add a vertical line at the selected date
        ax = self.figure.axes[0]
        self.vertical_line = ax.axvline(x=selected_date, color='red', linestyle='--', 
                                       linewidth=2, label='Selected Start Date')
        ax.legend()
        self.canvas.draw()
    
    def _enable_selection_mode(self, mode='start'):
        """Enable chart selection mode with appropriate settings
        
        Args:
            mode: 'start' for start date selection, 'qi' for qi selection
        """
        if self.figure is None:
            QMessageBox.warning(self, "Warning", "Please run an analysis first to display a chart.")
            return False
        
        # Disconnect any existing handlers
        if self.click_cid:
            self.figure.canvas.mpl_disconnect(self.click_cid)
            self.click_cid = None
        if self.qi_click_cid:
            self.figure.canvas.mpl_disconnect(self.qi_click_cid)
            self.qi_click_cid = None
        if self.key_cid:
            self.figure.canvas.mpl_disconnect(self.key_cid)
            self.key_cid = None
        
        # Store original cursor and change to crosshair
        self.original_cursor = self.canvas.cursor()
        self.canvas.setCursor(QCursor(Qt.CursorShape.CrossCursor))
        
        # Set focus to canvas to receive keyboard events
        self.canvas.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.canvas.setFocus()
        
        # Install event filter to catch ESC key (only if not already installed)
        if not self.event_filter_installed:
            self.canvas.installEventFilter(self)
            self.event_filter_installed = True
        
        # Connect the key press handler
        self.key_cid = self.figure.canvas.mpl_connect('key_press_event', self.on_key_press)
        
        return True

    # =========================================================================
    # END HELPER METHODS
    # =========================================================================

    def get_well_hierarchy(self, well_name):
        """Get the district, field, and formation for a given well"""
        df_well = self.df_all[self.df_all['Well_Name'] == well_name]
        if df_well.empty:
            return None, None, None
        district = df_well.iloc[0]['District']
        field = df_well.iloc[0]['Field']
        formation = df_well.iloc[0]['Alias_Formation']
        return str(district).strip(), str(field).strip(), str(formation).strip()
    
    def sanitize_filename(self, name):
        """Remove or replace invalid characters for Windows file/folder names"""
        # Invalid characters for Windows: < > : " / \ | ? *
        invalid_chars = '<>:"/\\|?*'
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        return sanitized
    
    def convert_to_json_serializable(self, obj):
        """Convert numpy/pandas types to JSON-serializable Python types"""
        if isinstance(obj, dict):
            return {key: self.convert_to_json_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self.convert_to_json_serializable(item) for item in obj]
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, (np.bool_, bool)):
            return bool(obj)
        elif pd.isna(obj):
            return None
        else:
            return obj
    
    def get_save_path(self, well_name):
        """Generate the save path for a well based on hierarchy"""
        district, field, formation = self.get_well_hierarchy(well_name)
        if not district or not field or not formation:
            return None
        
        # Sanitize all path components to remove invalid characters
        district_safe = self.sanitize_filename(district)
        field_safe = self.sanitize_filename(field)
        formation_safe = self.sanitize_filename(formation)
        well_safe = self.sanitize_filename(well_name)
        
        # Create path: Saved / District / Field / Formation / Well
        save_dir = Path("Saved") / district_safe / field_safe / formation_safe / well_safe
        save_dir.mkdir(parents=True, exist_ok=True)
        return save_dir
    
    def create_analysis_data(self, well_name):
        """Create a dictionary of all analysis data for saving"""
        if not self.current_analysis_params or self.current_well_name != well_name:
            return None
        
        # Get current GUI state
        # Get hierarchy for the well
        district, field, formation = self.get_well_hierarchy(well_name)
        
        data = {
            'timestamp': datetime.now().isoformat(),
            'well_name': well_name,
            'applied_wells': self.applied_wells if hasattr(self, 'applied_wells') else [well_name],  # Store all wells in selection
            'is_aggregated': len(self.applied_wells) > 1 if hasattr(self, 'applied_wells') else False,
            'district': district if district else '',
            'field': field if field else '',
            'formation': formation if formation else '',
            
            # Start method
            'start_method': 'auto' if self.auto_select_radio.isChecked() else 
                           ('manual' if self.manual_start_radio.isChecked() else 'qi'),
            'manual_year': self.manual_year_edit.text(),
            'manual_month': self.manual_month_edit.text(),
            'qi_year': self.qi_year_edit.text(),
            'qi_month': self.qi_month_edit.text(),
            'qi_value': self.qi_value_edit.text(),
            'fixed_qi': float(self.fixed_qi) if self.fixed_qi is not None else None,
            
            # Parameters
            'threshold': self.threshold_edit.text(),
            'threshold_opt': self.threshold_opt_check.isChecked(),
            'forecast_avg_points': self.forecast_avg_points_combo.currentText(),
            'forecast_duration': self.forecast_duration_edit.text(),
            
            # Auto start parameters
            'smooth_window': self.smooth_window_edit.text(),
            'smooth_window_opt': self.smooth_window_opt_check.isChecked(),
            'min_segment_length': self.min_segment_length_edit.text(),
            'min_segment_length_opt': self.min_segment_length_opt_check.isChecked(),
            'change_sensitivity': self.change_sensitivity_edit.text(),
            'change_sensitivity_opt': self.change_sensitivity_opt_check.isChecked(),
            'stability_window': self.stability_window_edit.text(),
            'stability_window_opt': self.stability_window_opt_check.isChecked(),
            'min_decline_rate': self.min_decline_rate_edit.text(),
            'min_decline_rate_opt': self.min_decline_rate_opt_check.isChecked(),
            
            # Plot options
            'show_outliers': self.show_outliers_check.isChecked(),
            'show_pre_decline': self.show_pre_decline_check.isChecked(),
            'show_forecast': self.show_forecast_check.isChecked(),
            'show_inliers': self.show_inliers_check.isChecked(),
            'show_channel': self.show_channel_check.isChecked(),
            
            # Optimization
            'optimize': self.optimize_check.isChecked(),
            
            # Active references for this analysis
            'active_references': self.get_active_references(),
            
            # Analysis results
            'results_text': self.results_text.toPlainText(),
            'analysis_params': self.convert_to_json_serializable(self.current_analysis_params)
        }
        
        # Convert all data to JSON-serializable types
        data = self.convert_to_json_serializable(data)
        
        return data
    
    def save_current_well(self):
        """Save analysis for currently selected well"""
        well_name = self.get_primary_well()
        if not well_name:
            QMessageBox.warning(self, "Warning", "Please select a well from the tree")
            return
        
        if not self.current_analysis_params or self.current_well_name != well_name:
            QMessageBox.warning(self, "Warning", "Please run analysis first before saving")
            return
        
        save_dir = self.get_save_path(well_name)
        if not save_dir:
            QMessageBox.critical(self, "Error", "Could not determine save path for this well")
            return
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"{timestamp}.json"
        filepath = save_dir / filename
        
        # Create analysis data
        data = self.create_analysis_data(well_name)
        if not data:
            QMessageBox.critical(self, "Error", "Could not create analysis data")
            return
        
        # Extract qi, Di, b from results_text and add to data
        results_text = self.results_text.toPlainText()
        qi_val = None
        di_val = None
        b_val = None
        
        for line in results_text.split('\n'):
            if 'qi =' in line and 'Di =' in line and 'b =' in line:
                try:
                    parts = line.split(',')
                    qi_val = float(parts[0].split('=')[1].strip())
                    di_val = float(parts[1].split('=')[1].strip())
                    b_val = float(parts[2].split('=')[1].strip())
                except:
                    pass
        
        if qi_val is not None:
            data['qi'] = qi_val
            data['Di'] = di_val
            data['b'] = b_val
        
        # Also get the actual start date from the decline phase (df_well after start_idx)
        well_name_in_data = self.current_well_name
        if well_name_in_data:
            df_temp = self.df_all[self.df_all['Well_Name'] == well_name_in_data].copy()
            if not df_temp.empty:
                df_temp = df_temp.sort_values('Prod_Date')
                if len(df_temp) > 0:
                    # Get the start date (first point in decline phase)
                    start_idx = self.current_analysis_params.get('manual_start_idx', 0)
                    if start_idx > 0:
                        df_temp = df_temp.iloc[start_idx:]
                    if len(df_temp) > 0:
                        data['decline_start_date'] = df_temp.iloc[0]['Prod_Date'].isoformat()
        
        # Save to file
        try:
            with open(filepath, 'w') as f:
                json.dump(data, f, indent=2)
            
            # Also store in session and mark as saved
            self.session_analyses[well_name] = data
            self.saved_analyses[well_name] = True
            
            # Update reference list
            self.update_reference_list()
            
            QMessageBox.information(self, "Success", 
                                   f"Analysis saved successfully!\n\nFile: {filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save analysis: {str(e)}")
    
    def get_saved_files_for_well(self, well_name):
        """Get all saved analysis files for a well, sorted by date (newest first)"""
        save_dir = self.get_save_path(well_name)
        if not save_dir or not save_dir.exists():
            return []
        
        # Get all .json files
        json_files = list(save_dir.glob("*.json"))
        
        # Sort by modification time, newest first
        json_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        return json_files
    
    def load_analysis_from_file(self, filepath):
        """Load analysis data from a JSON file"""
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
            return data
        except Exception as e:
            print(f"Error loading {filepath}: {e}")
            return None
    
    def apply_analysis_data(self, data):
        """Apply loaded analysis data to GUI"""
        if not data:
            return False
        
        try:
            # Set loading flag to prevent signal handlers
            self._loading_analysis = True
            
            # Set start method
            start_method = data.get('start_method', 'auto')
            if start_method == 'auto':
                self.auto_select_radio.setChecked(True)
            elif start_method == 'manual':
                self.manual_start_radio.setChecked(True)
            else:  # qi
                self.qi_radio.setChecked(True)
            
            # Set manual dates (with validation)
            self.manual_year_edit.setText(data.get('manual_year', '2020'))
            self.manual_month_edit.setText(data.get('manual_month', '1'))
            self.validate_manual_date()  # Validate and correct if needed
            
            self.qi_year_edit.setText(data.get('qi_year', '2020'))
            self.qi_month_edit.setText(data.get('qi_month', '1'))
            self.validate_qi_date()  # Validate and correct if needed
            
            self.qi_value_edit.setText(data.get('qi_value', ''))
            self.fixed_qi = data.get('fixed_qi')
            
            # Set parameters
            self.threshold_edit.setText(data.get('threshold', '2.0'))
            self.threshold_opt_check.setChecked(data.get('threshold_opt', True))
            self.forecast_avg_points_combo.setCurrentText(data.get('forecast_avg_points', 'The Model'))
            self.forecast_duration_edit.setText(data.get('forecast_duration', '60'))
            
            # Set auto start parameters
            self.smooth_window_edit.setText(data.get('smooth_window', '13'))
            self.smooth_window_opt_check.setChecked(data.get('smooth_window_opt', True))
            self.min_segment_length_edit.setText(data.get('min_segment_length', '12'))
            self.min_segment_length_opt_check.setChecked(data.get('min_segment_length_opt', True))
            self.change_sensitivity_edit.setText(data.get('change_sensitivity', '0.3'))
            self.change_sensitivity_opt_check.setChecked(data.get('change_sensitivity_opt', True))
            self.stability_window_edit.setText(data.get('stability_window', '6'))
            self.stability_window_opt_check.setChecked(data.get('stability_window_opt', True))
            self.min_decline_rate_edit.setText(data.get('min_decline_rate', '0.15'))
            self.min_decline_rate_opt_check.setChecked(data.get('min_decline_rate_opt', True))
            
            # Set plot options
            self.show_outliers_check.setChecked(data.get('show_outliers', True))
            self.show_pre_decline_check.setChecked(data.get('show_pre_decline', True))
            self.show_forecast_check.setChecked(data.get('show_forecast', True))
            self.show_inliers_check.setChecked(data.get('show_inliers', True))
            self.show_channel_check.setChecked(data.get('show_channel', False))
            
            # Set optimization
            self.optimize_check.setChecked(data.get('optimize', True))
            
            # Set results
            self.results_text.setText(data.get('results_text', ''))
            
            # Store analysis params
            self.current_analysis_params = data.get('analysis_params')
            self.current_well_name = data.get('well_name')
            
            # Restore applied wells selection (for aggregated analyses)
            applied_wells = data.get('applied_wells', [data.get('well_name')])
            if applied_wells:
                self.applied_wells = applied_wells
            
            # Store active references temporarily for restoration AFTER everything else is done
            # We'll restore them after the analysis is run and plot is generated
            active_references = data.get('active_references', [])
            self._pending_references_restore = active_references
            
            # Update reference list (but don't restore selection yet)
            try:
                self.update_reference_list()
            except Exception:
                pass
            
            return True
        except Exception as e:
            print(f"Error applying analysis data: {e}")
            return False
        finally:
            # Clear loading flag
            self._loading_analysis = False
    
    def load_current_well(self):
        """Show dialog to select and load an analysis for current well"""
        well_name = self.get_primary_well()
        if not well_name:
            QMessageBox.warning(self, "Warning", "Please select a well from the tree")
            return
        
        saved_files = self.get_saved_files_for_well(well_name)
        if not saved_files:
            QMessageBox.information(self, "No Saved Analyses", 
                                   f"No saved analyses found for well: {well_name}")
            return
        
        # Create selection dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Load Analysis - {well_name}")
        dialog.setMinimumSize(500, 400)
        layout = QVBoxLayout(dialog)
        
        label = QLabel("Select an analysis to load:")
        layout.addWidget(label)
        
        list_widget = QListWidget()
        for filepath in saved_files:
            # Parse timestamp from filename
            filename = filepath.stem  # Get filename without extension
            try:
                dt = datetime.strptime(filename, "%Y-%m-%d_%H-%M-%S")
                display_text = dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                display_text = filename
            
            list_widget.addItem(display_text)
        
        layout.addWidget(list_widget)
        
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | 
                                      QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            selected_idx = list_widget.currentRow()
            if selected_idx >= 0:
                filepath = saved_files[selected_idx]
                data = self.load_analysis_from_file(filepath)
                if data:
                    self.apply_analysis_data(data)
                    self.session_analyses[well_name] = data
                    self.saved_analyses[well_name] = True  # Loaded from file, so it's saved
                    
                    # Re-run analysis to display chart
                    QApplication.processEvents()
                    self.run_analysis()
                    
                    # Restore references AFTER everything else is done
                    self._restore_references_after_analysis()
                    
                    QMessageBox.information(self, "Success", 
                                           f"Analysis loaded successfully from:\n{filepath.name}")
                else:
                    QMessageBox.critical(self, "Error", "Failed to load analysis file")
    
    def load_latest_current_well(self):
        """Load the most recent analysis for current well"""
        well_name = self.get_primary_well()
        if not well_name:
            QMessageBox.warning(self, "Warning", "Please select a well from the tree")
            return
        
        saved_files = self.get_saved_files_for_well(well_name)
        if not saved_files:
            QMessageBox.information(self, "No Saved Analyses", 
                                   f"No saved analyses found for well: {well_name}")
            return
        
        # Load the first file (most recent)
        filepath = saved_files[0]
        data = self.load_analysis_from_file(filepath)
        if data:
            self.apply_analysis_data(data)
            self.session_analyses[well_name] = data
            self.saved_analyses[well_name] = True  # Loaded from file, so it's saved
            
            # Re-run analysis to display chart
            QApplication.processEvents()
            self.run_analysis()
            
            # Restore references AFTER everything else is done
            self._restore_references_after_analysis()
            
            QMessageBox.information(self, "Success", 
                                   f"Latest analysis loaded successfully!\n\nFile: {filepath.name}")
        else:
            QMessageBox.critical(self, "Error", "Failed to load analysis file")
    
    def load_all_latest(self):
        """Load the latest analysis for all wells that have saved analyses"""
        saved_root = Path("Saved")
        if not saved_root.exists():
            QMessageBox.information(self, "No Saved Analyses", 
                                   "No saved analyses folder found.")
            return
        
        loaded_count = 0
        failed_wells = []
        current_well_name = self.get_primary_well()
        current_well_data = None
        
        # Traverse the directory structure
        for district_dir in saved_root.iterdir():
            if not district_dir.is_dir():
                continue
            for field_dir in district_dir.iterdir():
                if not field_dir.is_dir():
                    continue
                for formation_dir in field_dir.iterdir():
                    if not formation_dir.is_dir():
                        continue
                    for well_dir in formation_dir.iterdir():
                        if not well_dir.is_dir():
                            continue
                        
                        # The well_dir.name is sanitized, need to match against actual well names
                        # Try to find the matching well in the data
                        well_name_sanitized = well_dir.name
                        
                        # Get all JSON files for this well
                        json_files = list(well_dir.glob("*.json"))
                        if not json_files:
                            continue
                        
                        # Sort by modification time, newest first
                        json_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                        
                        # Load the most recent
                        filepath = json_files[0]
                        data = self.load_analysis_from_file(filepath)
                        
                        if data:
                            # Get the actual well name from the data
                            actual_well_name = data.get('well_name', well_name_sanitized)
                            self.session_analyses[actual_well_name] = data
                            self.saved_analyses[actual_well_name] = True  # Loaded from file, so it's saved
                            loaded_count += 1
                            
                            # Check if this is the current well
                            if actual_well_name == current_well_name:
                                current_well_data = data
                        else:
                            failed_wells.append(well_name_sanitized)
        
        if loaded_count == 0:
            QMessageBox.information(self, "No Analyses Found", 
                                   "No saved analyses found in the Saved folder.")
            return
        
        # If current well has loaded data, apply it
        if current_well_data:
            self.apply_analysis_data(current_well_data)
            # Re-run analysis to display chart
            QApplication.processEvents()
            self.run_analysis()
            
            # Restore references AFTER everything else is done
            self._restore_references_after_analysis()
        
        if failed_wells:
            QMessageBox.warning(self, "Partial Success", 
                               f"Loaded {loaded_count} well(s) into session.\n\n"
                               f"Failed to load: {', '.join(failed_wells)}")
        else:
            message = f"Successfully loaded latest analyses for {loaded_count} well(s) into session!"
            if current_well_data:
                message += f"\n\nCurrent well ({current_well_name}) has been updated with its saved analysis."
            else:
                message += "\n\nSwitch to a well to view its loaded analysis."
            QMessageBox.information(self, "Success", message)
    
    def save_session(self):
        """Save all analyses in the current session to a single file"""
        if not self.session_analyses:
            QMessageBox.warning(self, "Warning", 
                               "No analyses in current session to save.\n\n"
                               "Please run analyses first or load existing analyses.")
            return
        
        # Ask user for filename
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Save Session",
            str(Path("Saved")),
            "Session Files (*.json);;All Files (*.*)"
        )
        
        if not filename:
            return  # User cancelled
        
        # Ensure the filename has .json extension
        if not filename.endswith('.json'):
            filename += '.json'
        
        try:
            # Save current analysis settings (plot options and references) to session before saving
            # This ensures the current analysis's state is preserved
            if self.current_well_name and self.current_well_name in self.session_analyses:
                self.save_current_settings_to_session()
            
            # Also save settings for all other analyses in the session
            # (in case they were modified but not saved)
            for well_name in self.session_analyses:
                if well_name != self.current_well_name:
                    # For non-current analyses, we keep the saved state
                    # If user wants to save changes to other analyses, they need to load them first
                    pass
            
            # Create session data structure
            session_data = {
                'timestamp': datetime.now().isoformat(),
                'version': '1.0',
                'well_count': len(self.session_analyses),
                'analyses': self.session_analyses
            }
            
            # Convert to JSON-serializable format
            session_data = self.convert_to_json_serializable(session_data)
            
            # Save to file
            with open(filename, 'w') as f:
                json.dump(session_data, f, indent=2)
            
            # Mark all wells as saved
            for well_name in self.session_analyses:
                self.saved_analyses[well_name] = True
            
            ref_count = 0
            QMessageBox.information(self, "Success", 
                                   f"Session saved successfully!\n\n"
                                   f"File: {Path(filename).name}\n"
                                   f"Number of Analyses: {len(self.session_analyses)}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save session: {str(e)}")
    
    def load_session(self):
        """Load all analyses from a session file"""
        # Open file dialog
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Load Session",
            str(Path("Saved")),
            "Session Files (*.json);;All Files (*.*)"
        )
        
        if not filename:
            return  # User cancelled
        
        try:
            # Load from file
            with open(filename, 'r') as f:
                session_data = json.load(f)
            
            # Check if it's a session file or old format
            if 'analyses' in session_data:
                # New session format
                analyses = session_data['analyses']
                well_count = session_data.get('well_count', len(analyses))
            elif 'well_name' in session_data:
                # Old format - single well analysis
                well_name = session_data['well_name']
                analyses = {well_name: session_data}
                well_count = 1
            else:
                QMessageBox.critical(self, "Error", "Invalid session file format")
                return
            
            # Check if there are unsaved analyses in current session
            unsaved_wells = []
            for well_name in self.session_analyses:
                if not self.saved_analyses.get(well_name, False):
                    unsaved_wells.append(well_name)
            
            if unsaved_wells:
                msg = QMessageBox(self)
                msg.setWindowTitle("Unsaved Changes")
                msg.setIcon(QMessageBox.Icon.Warning)
                msg.setText(f"You have {len(unsaved_wells)} unsaved analysis/analyses in the current session:")
                msg.setInformativeText("\n".join(unsaved_wells[:10]) + 
                                      ("\n..." if len(unsaved_wells) > 10 else "") +
                                      "\n\nLoading a new session will discard these changes.")
                
                continue_btn = msg.addButton("Continue Loading", QMessageBox.ButtonRole.AcceptRole)
                cancel_btn = msg.addButton("Cancel", QMessageBox.ButtonRole.RejectRole)
                msg.setDefaultButton(cancel_btn)
                
                msg.exec()
                
                if msg.clickedButton() == cancel_btn:
                    return
            
            # Clear current session and load new analyses
            self.session_analyses.clear()
            self.saved_analyses.clear()
            
            loaded_count = 0
            failed_wells = []
            
            for well_name, data in analyses.items():
                try:
                    self.session_analyses[well_name] = data
                    self.saved_analyses[well_name] = True  # Loaded from file, so it's saved
                    loaded_count += 1
                except Exception as e:
                    failed_wells.append(well_name)
                    print(f"Failed to load {well_name}: {e}")
            
            if loaded_count == 0:
                QMessageBox.warning(self, "Error", "Failed to load any analyses from the session file")
                return
            
            # Show success message without changing the current display
            if failed_wells:
                QMessageBox.warning(self, "Partial Success", 
                                   f"Loaded {loaded_count} well(s) into session.\n\n"
                                   f"Failed to load: {', '.join(failed_wells)}\n\n"
                                   f"Select wells from the tree and apply to view loaded analyses.")
            else:
                QMessageBox.information(self, "Success", 
                                       f"Successfully loaded {loaded_count} analysis/analyses into session!\n\n"
                                       f"Select wells from the tree and apply to view loaded analyses.")
            
            # Update reference list BEFORE showing the dialog
            # This prevents issues when the dialog closes and references are cleared
            self.update_reference_list()
            
            # Automatically show the analyses list and update export tab
            self.show_analyses_list()
            
            # Note: Don't call update_reference_list() again after dialog closes,
            # as it would clear references that were restored when loading an analysis
            # If an analysis was loaded from the dialog, its references are already restored
            
            # Update the analysis list in the export tab
            if hasattr(self, 'analysis_list'):
                self.update_analysis_list_for_export()
        
        except json.JSONDecodeError as e:
            QMessageBox.critical(self, "Error", f"Invalid JSON file: {str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load session: {str(e)}")
    
    def show_analyses_list(self):
        """Show a dialog with list of all analyses in the current session"""
        if not self.session_analyses:
            QMessageBox.information(self, "No Analyses", 
                                   "No analyses found in the current session.\n\n"
                                   "Run analyses on wells and they will appear here.")
            return
        
        # Create dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Available Analyses ({len(self.session_analyses)})")
        dialog.setMinimumSize(600, 400)
        layout = QVBoxLayout(dialog)
        
        # Add instruction label
        instruction = QLabel("Double-click on an analysis to load and display it:")
        instruction.setStyleSheet("font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        layout.addWidget(instruction)
        
        # Create list widget
        list_widget = QListWidget()
        list_widget.setStyleSheet("""
            QListWidget {
                font-size: 11px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:hover {
                background-color: #ecf0f1;
            }
            QListWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        
        # Populate list with analyses
        for well_name, data in self.session_analyses.items():
            applied_wells = data.get('applied_wells', [well_name])
            saved_status = "✓ Saved" if self.saved_analyses.get(well_name, False) else "✗ Unsaved"
            
            # Get meaningful display name based on hierarchy
            hierarchy_name = self._build_selection_title(applied_wells)
            
            # Format the display text
            display_text = f"{hierarchy_name}            {saved_status}"
            
            item = list_widget.addItem(display_text)
            # Store the well_name as item data for retrieval
            list_widget.item(list_widget.count() - 1).setData(Qt.ItemDataRole.UserRole, well_name)
        
        layout.addWidget(list_widget)
        
        # Add info label
        info_label = QLabel(f"Total: {len(self.session_analyses)} analysis/analyses in session")
        info_label.setStyleSheet("color: #7f8c8d; font-style: italic; margin-top: 5px;")
        layout.addWidget(info_label)
        
        # Action buttons: Load, Delete, Close
        button_layout = QHBoxLayout()
        load_btn = QPushButton("Load")
        load_btn.setEnabled(False)
        delete_btn = QPushButton("Delete")
        delete_btn.setEnabled(False)
        close_btn = QPushButton("Close")
        
        def update_action_buttons():
            has_selection = list_widget.currentItem() is not None
            load_btn.setEnabled(has_selection)
            delete_btn.setEnabled(has_selection)
        
        def on_load_clicked():
            item = list_widget.currentItem()
            if not item:
                return
            well_name = item.data(Qt.ItemDataRole.UserRole)
            if well_name and well_name in self.session_analyses:
                dialog.accept()
                self.load_analysis_from_session(well_name)
        
        def on_delete_clicked():
            item = list_widget.currentItem()
            if not item:
                return
            well_name = item.data(Qt.ItemDataRole.UserRole)
            if not well_name:
                return
            reply = QMessageBox.question(
                self,
                "Delete Analysis",
                f"Remove analysis for '{well_name}' from this session?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                # Check if this analysis is currently active in the chart
                is_active = well_name in self.applied_wells if self.applied_wells else False
                
                if well_name in self.session_analyses:
                    del self.session_analyses[well_name]
                if well_name in self.saved_analyses:
                    del self.saved_analyses[well_name]
                row = list_widget.row(item)
                list_widget.takeItem(row)
                info_label.setText(f"Total: {len(self.session_analyses)} analysis/analyses in session")
                update_action_buttons()
                
                # If the deleted analysis was active, reset chart and clear tree selection
                if is_active:
                    self.clear_selection()
                
                if list_widget.count() == 0:
                    dialog.accept()
        
        # Wire up events
        load_btn.clicked.connect(on_load_clicked)
        delete_btn.clicked.connect(on_delete_clicked)
        close_btn.clicked.connect(dialog.accept)
        list_widget.currentItemChanged.connect(lambda cur, prev: update_action_buttons())
        
        # Double-click also loads
        def on_item_double_clicked(item):
            list_widget.setCurrentItem(item)
            on_load_clicked()
        list_widget.itemDoubleClicked.connect(on_item_double_clicked)
        
        # Layout buttons
        button_layout.addWidget(load_btn)
        button_layout.addWidget(delete_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        # Show dialog
        dialog.exec()
    
    def load_analysis_from_session(self, well_name):
        """Load and display an analysis from the session by well name"""
        if well_name not in self.session_analyses:
            QMessageBox.warning(self, "Error", f"Analysis for '{well_name}' not found in session")
            return
        
        data = self.session_analyses[well_name]
        
        try:
            # Get the applied wells from the saved data
            applied_wells = data.get('applied_wells', [well_name])
            
            # Clear current selection
            self.clear_selection()
            
            # Select the wells in the tree
            self.select_wells_in_tree(applied_wells)
            
            # Apply the selection
            self.applied_wells = applied_wells
            
            # Update the status label
            if len(applied_wells) == 1:
                status_text = f"Applied: {applied_wells[0]}"
            else:
                status_text = f"Applied: {len(applied_wells)} wells"
            
            self.applied_selection_label.setText(status_text)
            self.applied_selection_label.setStyleSheet("""
                QLabel {
                    color: #27ae60;
                    font-weight: bold;
                    font-size: 10px;
                }
            """)
            
            # Apply the analysis data to GUI
            if self.apply_analysis_data(data):
                # Verify that applied_wells is properly set before running analysis
                if not self.applied_wells:
                    print(f"WARNING: applied_wells is empty before run_analysis")
                    
                # Now we need to re-run the analysis to generate the plot
                # Process events to ensure UI updates are complete before reading selections
                QApplication.processEvents()
                self.run_analysis()
                
                # Restore references AFTER everything else is done (plot generated, etc.)
                self._restore_references_after_analysis()
                
                # Process events once more to ensure references are restored
                QApplication.processEvents()
                
                # Show success message
                QMessageBox.information(self, "Success", 
                                       f"Successfully loaded analysis for: {well_name}")
            else:
                QMessageBox.warning(self, "Error", 
                                   "Failed to apply analysis data")
        except Exception as e:
            QMessageBox.critical(self, "Error", 
                               f"Failed to load analysis: {str(e)}")
    
    def select_wells_in_tree(self, well_names):
        """Select specified wells in the tree widget by checking their checkboxes"""
        if not well_names:
            return
        
        # Temporarily disconnect the signal to prevent triggering events
        self.well_tree.itemChanged.disconnect(self.on_tree_item_changed)
        self._updating_tree = True
        
        # Convert well_names to a set for faster lookup
        wells_to_select = set(well_names)
        
        # Traverse the tree and check matching wells
        def traverse_and_check(item):
            data = item.data(0, Qt.ItemDataRole.UserRole)
            if data and data['type'] == 'well' and data['name'] in wells_to_select:
                item.setCheckState(0, Qt.CheckState.Checked)
                # Update parent states
                parent = item.parent()
                while parent:
                    self._update_parent_state(parent)
                    parent = parent.parent()
            
            # Recursively check children
            for i in range(item.childCount()):
                traverse_and_check(item.child(i))
        
        # Traverse all top-level items
        for i in range(self.well_tree.topLevelItemCount()):
            traverse_and_check(self.well_tree.topLevelItem(i))
        
        self._updating_tree = False
        # Reconnect the signal
        self.well_tree.itemChanged.connect(self.on_tree_item_changed)
        
        # Refresh tree display
        self.refresh_tree_display()
    
    def exit_application(self):
        """Exit the application with unsaved changes check"""
        # Simply close the window, which will trigger closeEvent
        # This avoids showing the dialog twice
        self.close()
    
    def closeEvent(self, event):
        """Handle window close event (X button)"""
        # Check for unsaved analyses
        unsaved_wells = []
        for well_name in self.session_analyses:
            if not self.saved_analyses.get(well_name, False):
                unsaved_wells.append(well_name)
        
        if unsaved_wells:
            # Create custom message box with three buttons
            msg = QMessageBox(self)
            msg.setWindowTitle("Unsaved Changes")
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setText(f"You have {len(unsaved_wells)} unsaved analysis/analyses:")
            msg.setInformativeText("\n".join(unsaved_wells[:10]) + 
                                  ("\n..." if len(unsaved_wells) > 10 else ""))
            
            save_btn = msg.addButton("Save All && Exit", QMessageBox.ButtonRole.AcceptRole)
            dont_save_btn = msg.addButton("Exit Without Saving", QMessageBox.ButtonRole.DestructiveRole)
            cancel_btn = msg.addButton("Cancel", QMessageBox.ButtonRole.RejectRole)
            
            msg.setDefaultButton(save_btn)
            msg.exec()
            
            clicked_button = msg.clickedButton()
            
            if clicked_button == save_btn:
                # Save all and exit
                self.save_session()
                event.accept()
            elif clicked_button == dont_save_btn:
                # Exit without saving
                event.accept()
            else:
                # Cancel
                event.ignore()
        else:
            # No unsaved changes, just exit
            event.accept()
    
    def remove_vertical_line(self):
        """Remove the vertical line from the chart if it exists"""
        if self.vertical_line is not None and self.figure is not None:
            try:
                self.vertical_line.remove()
                self.canvas.draw()
            except:
                pass  # Line may have already been removed
            self.vertical_line = None
    
    def get_well_date_range(self):
        """Get the minimum and maximum dates for the current well"""
        if self.current_well_name is None:
            return None, None
        
        # Get the appropriate well data
        df_well = self._get_well_data_for_selection()
        if df_well is None or len(df_well) == 0:
            return None, None
        
        df_well['Prod_Date'] = pd.to_datetime(df_well['Prod_Date'])
        min_date = df_well['Prod_Date'].min()
        max_date = df_well['Prod_Date'].max()
        
        return min_date, max_date
    
    def validate_and_correct_date(self, year, month, is_qi=False):
        """Validate and correct date to be within well data range"""
        min_date, max_date = self.get_well_date_range()
        if min_date is None or max_date is None:
            return year, month  # Can't validate without data
        
        try:
            input_date = pd.Timestamp(year=int(year), month=int(month), day=1)
            
            # Clamp to valid range
            if input_date < min_date:
                return min_date.year, min_date.month
            elif input_date > max_date:
                return max_date.year, max_date.month
            else:
                return int(year), int(month)
        except (ValueError, TypeError):
            # Invalid input, return nearest valid date
            if min_date is not None:
                return min_date.year, min_date.month
            else:
                return int(year) if year.isdigit() else 2000, int(month) if month.isdigit() else 1
    
    def validate_manual_date(self):
        """Validate and correct manual date input"""
        try:
            year = self.manual_year_edit.text()
            month = self.manual_month_edit.text()
            corrected_year, corrected_month = self.validate_and_correct_date(year, month, is_qi=False)
            
            # Update if corrected
            if str(corrected_year) != year or str(corrected_month) != month:
                self.manual_year_edit.setText(str(corrected_year))
                self.manual_month_edit.setText(str(corrected_month))
        except:
            pass  # Silently ignore errors during typing
    
    def validate_qi_date(self):
        """Validate and correct qi date input"""
        try:
            year = self.qi_year_edit.text()
            month = self.qi_month_edit.text()
            corrected_year, corrected_month = self.validate_and_correct_date(year, month, is_qi=True)
            
            # Update if corrected
            if str(corrected_year) != year or str(corrected_month) != month:
                self.qi_year_edit.setText(str(corrected_year))
                self.qi_month_edit.setText(str(corrected_month))
        except:
            pass  # Silently ignore errors during typing
    
    def update_vertical_line_from_inputs(self):
        """Add or update vertical line based on current manual date inputs"""
        if self.figure is None or self.canvas is None:
            return
        
        # Only show vertical line if in manual mode
        if self.auto_select_radio.isChecked():
            self.remove_vertical_line()
            return
        
        try:
            # Get the appropriate year and month based on which manual mode is selected
            if self.manual_start_radio.isChecked():
                year = int(self.manual_year_edit.text())
                month = int(self.manual_month_edit.text())
            elif self.qi_radio.isChecked():
                year = int(self.qi_year_edit.text())
                month = int(self.qi_month_edit.text())
            else:
                return
            
            # Create the date from inputs
            selected_date = pd.to_datetime(f"{year}-{month:02d}-01")
            
            # Remove existing line
            self.remove_vertical_line()
            
            # Add new vertical line at the selected date
            ax = self.figure.axes[0]
            self.vertical_line = ax.axvline(x=selected_date, color='red', linestyle='--', linewidth=2, label='Selected Start Date')
            ax.legend()
            self.canvas.draw()
        except (ValueError, AttributeError):
            # Invalid date input, don't draw line
            pass
    
    def eventFilter(self, obj, event):
        """Event filter to catch ESC key press globally"""
        if event.type() == QEvent.Type.KeyPress:
            if event.key() == Qt.Key.Key_Escape:
                if self.click_selection_enabled or self.qi_selection_enabled:
                    self.disable_selection_mode()
                    return True
        return super().eventFilter(obj, event)

    def enable_chart_click_selection(self):
        if not self._enable_selection_mode('start'):
            return
        
        self.click_selection_enabled = True
        self.qi_selection_enabled = False
        self.click_cid = self.figure.canvas.mpl_connect('button_press_event', self.on_chart_click)

    def on_chart_click(self, event):
        if not self.click_selection_enabled or event.inaxes is None:
            return
        clicked_date = event.xdata
        if np.isnan(clicked_date):
            return
        date_clicked = pd.to_datetime(clicked_date, unit='D')
        
        # Get the appropriate well data
        df_well = self._get_well_data_for_selection()
        if df_well is None:
            return
        
        df_well['Prod_Date'] = pd.to_datetime(df_well['Prod_Date'])
        min_date = df_well['Prod_Date'].min()
        max_date = df_well['Prod_Date'].max()
        
        # Validate date is within range
        if date_clicked < min_date:
            date_clicked = min_date
        elif date_clicked > max_date:
            date_clicked = max_date
        
        df_well = df_well.sort_values('Prod_Date')
        idx = np.searchsorted(df_well['Prod_Date'], date_clicked)
        if idx >= len(df_well):
            idx = len(df_well) - 1
        elif idx < 0:
            idx = 0
        selected_date = df_well.iloc[idx]['Prod_Date']
        
        # Update manual date inputs
        self.manual_year_edit.setText(str(selected_date.year))
        self.manual_month_edit.setText(str(selected_date.month))
        self.manual_start_radio.setChecked(True)
        
        # Draw vertical line at selected date
        self._draw_vertical_line_at_date(selected_date)
        
        self.disable_selection_mode()

    def enable_qi_click_selection(self):
        if not self._enable_selection_mode('qi'):
            return
        
        self.qi_selection_enabled = True
        self.click_selection_enabled = False
        self.qi_click_cid = self.figure.canvas.mpl_connect('button_press_event', self.on_qi_chart_click)

    def on_qi_chart_click(self, event):
        if not self.qi_selection_enabled or event.inaxes is None:
            return
        clicked_date = event.xdata
        clicked_rate = event.ydata
        if np.isnan(clicked_date) or np.isnan(clicked_rate) or clicked_rate <= 0:
            QMessageBox.warning(self, "Invalid Point", "Please click on a valid data point with positive rate.")
            return
        date_clicked = pd.to_datetime(clicked_date, unit='D')
        
        # Get the appropriate well data
        df_well = self._get_well_data_for_selection()
        if df_well is None:
            return
        
        df_well['Prod_Date'] = pd.to_datetime(df_well['Prod_Date'])
        min_date = df_well['Prod_Date'].min()
        max_date = df_well['Prod_Date'].max()
        
        # Validate date is within range
        if date_clicked < min_date:
            date_clicked = min_date
        elif date_clicked > max_date:
            date_clicked = max_date
        
        df_well = df_well.sort_values('Prod_Date')
        idx = np.searchsorted(df_well['Prod_Date'], date_clicked)
        if idx >= len(df_well):
            idx = len(df_well) - 1
        elif idx < 0:
            idx = 0
        selected_date = df_well.iloc[idx]['Prod_Date']
        
        # Update Qi inputs
        self.qi_year_edit.setText(str(selected_date.year))
        self.qi_month_edit.setText(str(selected_date.month))
        self.fixed_qi = float(clicked_rate)
        self.qi_value_edit.setText(f"{clicked_rate:.2f}")
        self.qi_radio.setChecked(True)
        
        # Draw vertical line at selected date
        self._draw_vertical_line_at_date(selected_date)
        
        self.disable_selection_mode()
    
    def on_key_press(self, event):
        """Handle key press events, particularly ESC to cancel selection"""
        if event.key == 'escape':
            self.disable_selection_mode()
    
    def disable_selection_mode(self):
        """Disable selection mode and restore original cursor"""
        self.click_selection_enabled = False
        self.qi_selection_enabled = False
        
        # Disconnect all event handlers
        if self.click_cid:
            self.figure.canvas.mpl_disconnect(self.click_cid)
            self.click_cid = None
        if self.qi_click_cid:
            self.figure.canvas.mpl_disconnect(self.qi_click_cid)
            self.qi_click_cid = None
        if self.key_cid:
            self.figure.canvas.mpl_disconnect(self.key_cid)
            self.key_cid = None
        
        # Note: We don't remove the event filter here anymore
        # It stays installed on the canvas and just checks the flags
        # This way it works for subsequent clicks
        
        # Restore original cursor
        if self.canvas and self.original_cursor is not None:
            self.canvas.setCursor(self.original_cursor)
        elif self.canvas:
            self.canvas.setCursor(QCursor(Qt.CursorShape.ArrowCursor))

    def create_general_tab(self):
        general_tab = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setSpacing(5)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        general_tab.setLayout(main_layout)
        
        label_width = 180  # Slightly wider for longer labels in General tab
        checkbox_width = 24  # Fixed width for checkbox area to ensure alignment
        
        # Row 1 - Outlier Threshold (with checkbox)
        row1 = QHBoxLayout()
        row1.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.threshold_opt_check = QCheckBox()
        self.threshold_opt_check.setChecked(True)
        self.threshold_opt_check.setFixedWidth(checkbox_width)
        row1.addWidget(self.threshold_opt_check)
        label1 = QLabel("Outlier Threshold:")
        label1.setFixedWidth(label_width)
        row1.addWidget(label1)
        self.threshold_edit = QLineEdit("2.0")
        self.threshold_edit.setMaximumWidth(100)
        row1.addWidget(self.threshold_edit)
        row1.addStretch()
        main_layout.addLayout(row1)
        
        # Row 2 - Forecast Avg Points (no checkbox, add spacer)
        row2 = QHBoxLayout()
        row2.setAlignment(Qt.AlignmentFlag.AlignLeft)
        spacer2 = QLabel("")  # Spacer to align with checkbox width
        spacer2.setFixedWidth(checkbox_width)
        row2.addWidget(spacer2)
        label2 = QLabel("Forecast Avg Points:")
        label2.setFixedWidth(label_width)
        row2.addWidget(label2)
        self.forecast_avg_points_combo = QComboBox()
        self.forecast_avg_points_combo.addItems(["The Model", "The Last Rate", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"])
        self.forecast_avg_points_combo.setCurrentText("The Model")
        self.forecast_avg_points_combo.setMinimumWidth(100)
        self.forecast_avg_points_combo.currentTextChanged.connect(self.on_forecast_avg_points_changed)
        row2.addWidget(self.forecast_avg_points_combo)
        row2.addStretch()
        main_layout.addLayout(row2)
        
        # Row 3 - Forecast Duration
        row3 = QHBoxLayout()
        row3.setAlignment(Qt.AlignmentFlag.AlignLeft)
        spacer3 = QLabel("")
        spacer3.setFixedWidth(checkbox_width)
        row3.addWidget(spacer3)
        label3 = QLabel("Forecast Duration (months):")
        label3.setFixedWidth(label_width)
        row3.addWidget(label3)
        self.forecast_duration_edit = QLineEdit("60")
        self.forecast_duration_edit.setMaximumWidth(100)
        row3.addWidget(self.forecast_duration_edit)
        row3.addStretch()
        main_layout.addLayout(row3)
        
        main_layout.addStretch()
        self.settings_tabs.addTab(general_tab, "General")

    def create_auto_start_tab(self):
        auto_start_tab = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setSpacing(5)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        auto_start_tab.setLayout(main_layout)
        
        label_width = 180  # Fixed width for labels to align input fields
        
        # Row 1
        row1 = QHBoxLayout()
        row1.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.smooth_window_opt_check = QCheckBox()
        self.smooth_window_opt_check.setChecked(True)
        row1.addWidget(self.smooth_window_opt_check)
        label1 = QLabel("Smooth Window:")
        label1.setFixedWidth(label_width)
        row1.addWidget(label1)
        self.smooth_window_edit = QLineEdit("13")
        self.smooth_window_edit.setMaximumWidth(100)
        row1.addWidget(self.smooth_window_edit)
        row1.addStretch()
        main_layout.addLayout(row1)
        
        # Row 2
        row2 = QHBoxLayout()
        row2.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.min_segment_length_opt_check = QCheckBox()
        self.min_segment_length_opt_check.setChecked(True)
        row2.addWidget(self.min_segment_length_opt_check)
        label2 = QLabel("Min Segment Length:")
        label2.setFixedWidth(label_width)
        row2.addWidget(label2)
        self.min_segment_length_edit = QLineEdit("12")
        self.min_segment_length_edit.setMaximumWidth(100)
        row2.addWidget(self.min_segment_length_edit)
        row2.addStretch()
        main_layout.addLayout(row2)
        
        # Row 3
        row3 = QHBoxLayout()
        row3.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.change_sensitivity_opt_check = QCheckBox()
        self.change_sensitivity_opt_check.setChecked(True)
        row3.addWidget(self.change_sensitivity_opt_check)
        label3 = QLabel("Change Sensitivity:")
        label3.setFixedWidth(label_width)
        row3.addWidget(label3)
        self.change_sensitivity_edit = QLineEdit("0.3")
        self.change_sensitivity_edit.setMaximumWidth(100)
        row3.addWidget(self.change_sensitivity_edit)
        row3.addStretch()
        main_layout.addLayout(row3)
        
        # Row 4
        row4 = QHBoxLayout()
        row4.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.stability_window_opt_check = QCheckBox()
        self.stability_window_opt_check.setChecked(True)
        row4.addWidget(self.stability_window_opt_check)
        label4 = QLabel("Stability Window:")
        label4.setFixedWidth(label_width)
        row4.addWidget(label4)
        self.stability_window_edit = QLineEdit("6")
        self.stability_window_edit.setMaximumWidth(100)
        row4.addWidget(self.stability_window_edit)
        row4.addStretch()
        main_layout.addLayout(row4)
        
        # Row 5
        row5 = QHBoxLayout()
        row5.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.min_decline_rate_opt_check = QCheckBox()
        self.min_decline_rate_opt_check.setChecked(True)
        row5.addWidget(self.min_decline_rate_opt_check)
        label5 = QLabel("Min Decline Rate:")
        label5.setFixedWidth(label_width)
        row5.addWidget(label5)
        self.min_decline_rate_edit = QLineEdit("0.15")
        self.min_decline_rate_edit.setMaximumWidth(100)
        row5.addWidget(self.min_decline_rate_edit)
        row5.addStretch()
        main_layout.addLayout(row5)
        
        main_layout.addStretch()
        self.settings_tabs.addTab(auto_start_tab, "Auto Start Detection")

    def create_plot_options_tab(self):
        plot_tab = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(6)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        plot_tab.setLayout(layout)
        
        self.show_outliers_check = QCheckBox("Show Outliers")
        self.show_outliers_check.setChecked(True)
        self.show_outliers_check.stateChanged.connect(self.update_plot_options)
        layout.addWidget(self.show_outliers_check)
        
        self.show_inliers_check = QCheckBox("Show Inliers Points")
        self.show_inliers_check.setChecked(True)
        self.show_inliers_check.stateChanged.connect(self.update_plot_options)
        layout.addWidget(self.show_inliers_check)
        
        self.show_pre_decline_check = QCheckBox("Show Pre-Decline Points")
        self.show_pre_decline_check.setChecked(True)
        self.show_pre_decline_check.stateChanged.connect(self.update_plot_options)
        layout.addWidget(self.show_pre_decline_check)
        
        self.show_forecast_check = QCheckBox("Show Forecast")
        self.show_forecast_check.setChecked(True)
        self.show_forecast_check.stateChanged.connect(self.update_plot_options)
        layout.addWidget(self.show_forecast_check)
        
        self.show_channel_check = QCheckBox("Show Channel")
        self.show_channel_check.setChecked(False)
        self.show_channel_check.stateChanged.connect(self.update_plot_options)
        layout.addWidget(self.show_channel_check)
        
        layout.addStretch()
        self.settings_tabs.addTab(plot_tab, "Plot Options")

    def create_exports_tab(self):
        """Create the Exports tab with language selection and export options"""
        exports_tab = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        layout.setContentsMargins(20, 20, 20, 20)
        exports_tab.setLayout(layout)
        
        # Analysis selection group
        analysis_group = QGroupBox("Analysis Selection")
        analysis_layout = QVBoxLayout()
        
        self.analysis_list = QListWidget()
        self.analysis_list.setMaximumHeight(150)
        self.analysis_list.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection)
        # Apply better selection styling
        self.analysis_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                background-color: white;
            }
            QListWidget::item {
                padding: 5px;
                border-bottom: 1px solid #e0e0e0;
            }
            QListWidget::item:selected {
                background-color: #4A90E2;
                color: white;
                border: none;
            }
            QListWidget::item:hover {
                background-color: #f0f0f0;
            }
            QListWidget::item:selected:hover {
                background-color: #3a7bc8;
            }
        """)
        analysis_layout.addWidget(QLabel("Select analyses to export:"))
        analysis_layout.addWidget(self.analysis_list)
        
        # Select buttons
        select_buttons_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("Select All")
        self.select_all_btn.clicked.connect(self.select_all_analyses)
        self.select_all_btn.setProperty("class", "secondary")
        self.deselect_all_btn = QPushButton("Deselect All")
        self.deselect_all_btn.clicked.connect(self.deselect_all_analyses)
        self.deselect_all_btn.setProperty("class", "secondary")
        select_buttons_layout.addWidget(self.select_all_btn)
        select_buttons_layout.addWidget(self.deselect_all_btn)
        analysis_layout.addLayout(select_buttons_layout)
        
        analysis_group.setLayout(analysis_layout)
        layout.addWidget(analysis_group)
        
        # Export language group
        language_group = QGroupBox("Export Language")
        language_layout = QVBoxLayout()
        
        self.export_language_group = QButtonGroup()
        self.english_radio = QRadioButton("English")
        self.english_radio.setChecked(True)
        self.persian_radio = QRadioButton("Persian")
        self.export_language_group.addButton(self.english_radio, 0)
        self.export_language_group.addButton(self.persian_radio, 1)
        
        language_layout.addWidget(self.english_radio)
        language_layout.addWidget(self.persian_radio)
        
        language_group.setLayout(language_layout)
        layout.addWidget(language_group)
        
        # Export duration group
        duration_group = QGroupBox("Export Duration")
        duration_layout = QVBoxLayout()
        
        self.export_history_check = QCheckBox("History")
        self.export_history_check.setChecked(True)
        duration_layout.addWidget(self.export_history_check)
        
        self.export_forecast_check = QCheckBox("Forecast")
        self.export_forecast_check.setChecked(True)
        duration_layout.addWidget(self.export_forecast_check)
        
        duration_group.setLayout(duration_layout)
        layout.addWidget(duration_group)
        
        # Export frequency group
        frequency_group = QGroupBox("Export Frequency")
        frequency_layout = QVBoxLayout()
        
        self.export_frequency_group = QButtonGroup()
        self.monthly_radio = QRadioButton("Monthly")
        self.monthly_radio.setChecked(True)
        self.yearly_radio = QRadioButton("Yearly")
        self.export_frequency_group.addButton(self.monthly_radio, 0)
        self.export_frequency_group.addButton(self.yearly_radio, 1)
        
        frequency_layout.addWidget(self.monthly_radio)
        frequency_layout.addWidget(self.yearly_radio)
        
        frequency_group.setLayout(frequency_layout)
        layout.addWidget(frequency_group)
        
        # Forecast Start Date group
        fsd_group = QGroupBox("Forecast Start Date")
        fsd_layout = QVBoxLayout()
        
        self.fsd_mode_group = QButtonGroup()
        self.fsd_auto_radio = QRadioButton("Auto (month after last history for each analysis)")
        self.fsd_auto_radio.setChecked(True)
        self.fsd_fixed_radio = QRadioButton("Fixed date (same for all analyses)")
        self.fsd_mode_group.addButton(self.fsd_auto_radio, 0)
        self.fsd_mode_group.addButton(self.fsd_fixed_radio, 1)
        
        fsd_layout.addWidget(self.fsd_auto_radio)
        
        # Fixed date layout
        fixed_date_layout = QHBoxLayout()
        fixed_date_layout.addWidget(self.fsd_fixed_radio)
        
        # Add date input (disabled by default)
        self.fsd_date_edit = QLineEdit()
        self.fsd_date_edit.setPlaceholderText("YYYY-MM-DD")
        self.fsd_date_edit.setEnabled(False)
        self.fsd_date_edit.setMaximumWidth(150)
        fixed_date_layout.addWidget(self.fsd_date_edit)
        fixed_date_layout.addStretch()
        
        fsd_layout.addLayout(fixed_date_layout)
        
        # Connect radio button to enable/disable date edit
        self.fsd_fixed_radio.toggled.connect(self.fsd_date_edit.setEnabled)
        
        fsd_group.setLayout(fsd_layout)
        layout.addWidget(fsd_group)
        
        # Include summary checkbox
        self.include_summary_check = QCheckBox("Include Summary Page")
        self.include_summary_check.setChecked(True)
        layout.addWidget(self.include_summary_check)
        
        # Export and Cancel buttons
        button_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("Export")
        self.export_btn.setProperty("class", "primary")
        self.export_btn.clicked.connect(self.export_to_excel)
        button_layout.addWidget(self.export_btn)
        
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setProperty("class", "secondary")
        self.cancel_btn.clicked.connect(lambda: self.main_tabs.setCurrentIndex(0))
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        # Add to main tabs
        self.main_tabs.addTab(exports_tab, "Exports")
        
        # Update the analysis list when tab is shown
        self.main_tabs.currentChanged.connect(self.update_analysis_list)

    def create_reference_tab(self):
        reference_tab = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(6)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        reference_tab.setLayout(layout)
        
        # Label for reference selection
        layout.addWidget(QLabel("Select Reference Analyses:"))
        
        # Create list widget for multiple reference selection
        self.reference_list = QListWidget()
        self.reference_list.setMaximumHeight(200)
        self.reference_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        self.reference_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #dee2e6;
                border-radius: 4px;
                background-color: white;
                font-size: 11px;
            }
            QListWidget::item {
                padding: 4px 8px;
                border-bottom: 1px solid #f8f9fa;
            }
            QListWidget::item:selected {
                background-color: #007bff;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #e9ecef;
            }
        """)
        layout.addWidget(self.reference_list)
        
        # Button layout for Select All and Deselect All
        button_layout = QHBoxLayout()
        
        self.select_all_btn = QPushButton("Select All")
        self.select_all_btn.clicked.connect(self.select_all_references)
        button_layout.addWidget(self.select_all_btn)
        
        self.deselect_all_btn = QPushButton("Deselect All")
        self.deselect_all_btn.clicked.connect(self.deselect_all_references)
        button_layout.addWidget(self.deselect_all_btn)
        
        layout.addLayout(button_layout)
        
        # Update reference list when session changes
        self.reference_list.itemSelectionChanged.connect(self.update_reference_display)
        
        layout.addStretch()
        self.settings_tabs.addTab(reference_tab, "Reference")

    def update_reference_list(self):
        """Update the reference list with all available analyses in the session"""
        self.reference_list.clear()
        
        if not self.session_analyses:
            return
        
        # Add all analyses from session to the reference list (excluding current active analysis)
        for well_name, data in sorted(self.session_analyses.items()):
            # Skip the current active analysis
            if well_name == self.current_well_name:
                continue
                
            # Create display name with hierarchy information
            hierarchy_parts = well_name.split('\\')
            if len(hierarchy_parts) >= 3:
                hierarchy_name = f"{hierarchy_parts[-3]}\\{hierarchy_parts[-2]}\\{hierarchy_parts[-1]}"
            else:
                hierarchy_name = well_name
            
            # Add analysis info if available
            if 'qi' in data and 'Di' in data and 'b' in data:
                qi = data['qi']
                Di = data['Di']
                b = data['b']
                display_text = f"{hierarchy_name} (qi={qi:.1f}, Di={Di:.4f}, b={b:.2f})"
            else:
                display_text = f"{hierarchy_name} (No fitted parameters)"
            
            item = QListWidgetItem(display_text)
            item.setData(Qt.ItemDataRole.UserRole, well_name)  # Store the actual well name
            self.reference_list.addItem(item)
    
    def select_all_references(self):
        """Select all items in the reference list"""
        # Skip if we're currently loading analysis data
        if hasattr(self, '_loading_analysis') and self._loading_analysis:
            return
        
        for i in range(self.reference_list.count()):
            self.reference_list.item(i).setSelected(True)
        
        # Save to session immediately
        self.save_current_settings_to_session()
    
    def deselect_all_references(self):
        """Deselect all items in the reference list"""
        # Skip if we're currently loading analysis data
        if hasattr(self, '_loading_analysis') and self._loading_analysis:
            return
        
        for i in range(self.reference_list.count()):
            self.reference_list.item(i).setSelected(False)
        
        # Save to session immediately
        self.save_current_settings_to_session()
    
    def get_active_references(self):
        """Get list of currently selected reference well names"""
        active_refs = []
        for i in range(self.reference_list.count()):
            item = self.reference_list.item(i)
            if item.isSelected():
                well_name = item.data(Qt.ItemDataRole.UserRole)
                if well_name and well_name in self.session_analyses:
                    # Get reference parameters from session
                    ref_data = self.session_analyses[well_name]
                    active_refs.append({
                        'well_name': well_name,
                        'qi': ref_data.get('qi'),
                        'Di': ref_data.get('Di'),
                        'b': ref_data.get('b')
                    })
        return active_refs
    
    def save_current_settings_to_session(self):
        """Save current plot options and references to session data"""
        if self.current_well_name and self.current_well_name in self.session_analyses:
            session_data = self.session_analyses[self.current_well_name]
            
            # Update plot options
            session_data['show_outliers'] = self.show_outliers_check.isChecked()
            session_data['show_pre_decline'] = self.show_pre_decline_check.isChecked()
            session_data['show_forecast'] = self.show_forecast_check.isChecked()
            session_data['show_inliers'] = self.show_inliers_check.isChecked()
            session_data['show_channel'] = self.show_channel_check.isChecked()
            
            # Update active references
            session_data['active_references'] = self.get_active_references()
            
            # Mark as unsaved if not already saved to file
            if hasattr(self, 'saved_analyses'):
                # Only mark as unsaved if it was previously saved (had file on disk)
                # Don't change from True to False if it's already False
                pass
    
    def restore_active_references(self, active_references):
        """Restore active references selection from saved data"""
        # Handle None, empty list, or invalid format
        if not active_references:
            self.deselect_all_references()
            return
        
        # Ensure active_references is a list
        if not isinstance(active_references, list):
            self.deselect_all_references()
            return
        
        # Deselect all first
        self.deselect_all_references()
        
        # Extract well names - handle both dict format and string format (backward compatibility)
        well_names_to_select = []
        for ref in active_references:
            if isinstance(ref, dict):
                well_name = ref.get('well_name')
            elif isinstance(ref, str):
                well_name = ref
            else:
                continue
            
            if well_name:
                well_names_to_select.append(well_name)
        
        # Select the matching items
        selected_count = 0
        for i in range(self.reference_list.count()):
            item = self.reference_list.item(i)
            if item is None:
                continue
                
            well_name = item.data(Qt.ItemDataRole.UserRole)
            
            if well_name and well_name in well_names_to_select:
                item.setSelected(True)
                selected_count += 1
        
        # Update the display to show selections
        self.reference_list.viewport().update()
        
        # Trigger plot update if we have an active analysis (references affect the plot)
        if self.current_analysis_params and self.current_well_name:
            # Use a small delay to ensure everything is set before updating plot
            QTimer.singleShot(50, self.update_reference_display)
        
        # If we didn't restore any but expected some, it might be because the list hasn't been updated yet
        # or the well names don't match. Try again after a delay
        if selected_count == 0 and well_names_to_select:
            # Try once more after a small delay to allow the list to be fully populated
            QTimer.singleShot(150, lambda: self._retry_restore_references(well_names_to_select))
    
    def _retry_restore_references(self, well_names_to_select):
        """Retry restoring references after a delay - called if initial restore failed"""
        # Update reference list to ensure it's current
        self.update_reference_list()
        
        # Try to select again
        for i in range(self.reference_list.count()):
            item = self.reference_list.item(i)
            if item is None:
                continue
                
            well_name = item.data(Qt.ItemDataRole.UserRole)
            if well_name and well_name in well_names_to_select:
                item.setSelected(True)
        
        # Update the display
        self.reference_list.viewport().update()
        # Trigger plot update if we have an active analysis
        if self.current_analysis_params and self.current_well_name:
            self.update_reference_display()
    
    def _final_restore_references(self, active_references):
        """Final step to restore references after everything else is done"""
        # Ensure we're not in loading state
        if hasattr(self, '_loading_analysis') and self._loading_analysis:
            # If still loading, try again later
            QTimer.singleShot(200, lambda: self._final_restore_references(active_references))
            return
        
        # Don't call update_reference_list() here - it was already called in 
        # _restore_references_after_analysis() and would clear selections
        # Just restore the references directly
        self.restore_active_references(active_references)
    
    def _restore_references_after_analysis(self):
        """Helper to restore pending references after analysis runs"""
        if hasattr(self, '_pending_references_restore'):
            active_references = self._pending_references_restore
            delattr(self, '_pending_references_restore')
            
            # Ensure reference list is up to date first
            self.update_reference_list()
            
            # Process events to ensure the list is fully built
            QApplication.processEvents()
            
            # Restore references synchronously - the list should be ready now
            # Only use delay if we're still in loading state
            if hasattr(self, '_loading_analysis') and self._loading_analysis:
                QTimer.singleShot(100, lambda: self._final_restore_references(active_references))
            else:
                self._final_restore_references(active_references)
    
    def update_reference_display(self):
        """Update the plot when reference selection or checkbox changes"""
        # Check if we have an analysis to update
        if self.current_analysis_params is None or self.current_well_name is None:
            return
        
        # Skip if we're currently loading analysis data (to avoid overwriting during restore)
        if hasattr(self, '_loading_analysis') and self._loading_analysis:
            return
        
        # Save current references to session immediately
        self.save_current_settings_to_session()
        
        # Trigger plot update (similar to update_plot_options)
        if self.figure is not None:
            self.update_plot_options()

    def on_forecast_avg_points_changed(self):
        """Handle forecast avg points combo box change - update chart immediately"""
        if self.current_analysis_params is None or self.current_well_name is None:
            return
        
        # Skip if we're currently loading analysis data
        if self._loading_analysis:
            return
        
        # Update current_analysis_params with new forecast_avg_points value
        forecast_avg_text = self.forecast_avg_points_combo.currentText()
        if forecast_avg_text == "The Model":
            forecast_avg_points = 0
        elif forecast_avg_text == "The Last Rate":
            forecast_avg_points = 1
        else:
            forecast_avg_points = int(forecast_avg_text)
        
        # Update the current analysis params
        self.current_analysis_params['forecast_avg_points'] = forecast_avg_points
        
        # Update the chart immediately
        self.update_plot_options()
        
        # Update session data if exists
        if hasattr(self, 'current_well_name') and self.current_well_name:
            self.update_current_session_forecast_avg_points(forecast_avg_points)

    def update_current_session_forecast_avg_points(self, forecast_avg_points):
        """Update forecast_avg_points in current session data"""
        if not hasattr(self, 'current_well_name') or not self.current_well_name:
            return
        
        # Convert forecast_avg_points back to text for storage
        if forecast_avg_points == 0:
            forecast_avg_text = "The Model"
        elif forecast_avg_points == 1:
            forecast_avg_text = "The Last Rate"
        else:
            forecast_avg_text = str(forecast_avg_points)
        
        # Update in session_analyses if exists
        if hasattr(self, 'session_analyses') and self.current_well_name in self.session_analyses:
            session_data = self.session_analyses[self.current_well_name]
            
            # Update in analysis_params
            if 'analysis_params' in session_data:
                session_data['analysis_params']['forecast_avg_points'] = forecast_avg_text
            
            # Update in analysis data
            if 'analysis' in session_data:
                session_data['analysis']['forecast_avg_points'] = forecast_avg_text
            
            # Update the main forecast_avg_points field
            session_data['forecast_avg_points'] = forecast_avg_text
            
            # Mark as unsaved since we modified the session
            if hasattr(self, 'saved_analyses'):
                self.saved_analyses[self.current_well_name] = False

    def update_plot_options(self):
        # Check if we have an analysis to update
        if self.current_analysis_params is None or self.current_well_name is None:
            return
        if self.figure is None:
            return
        
        try:
            show_outliers = self.show_outliers_check.isChecked()
            show_pre_decline = self.show_pre_decline_check.isChecked()
            show_forecast = self.show_forecast_check.isChecked()
            show_inliers = self.show_inliers_check.isChecked()
            show_channel = self.show_channel_check.isChecked()
            
            # Get reference model parameters
            reference_models = []
            selected_items = self.reference_list.selectedItems()
            for item in selected_items:
                well_name = item.data(Qt.ItemDataRole.UserRole)
                if well_name in self.session_analyses:
                    data = self.session_analyses[well_name]
                    if 'qi' in data and 'Di' in data and 'b' in data:
                        qi = data['qi']
                        Di = data['Di']
                        b = data['b']
                        # Create reference model dict with well name for display
                        hierarchy_parts = well_name.split('\\')
                        if len(hierarchy_parts) >= 3:
                            display_name = f"{hierarchy_parts[-3]}\\{hierarchy_parts[-2]}\\{hierarchy_parts[-1]}"
                        else:
                            display_name = well_name
                        
                        reference_models.append({
                            'Di': Di, 
                            'b': b, 
                            'name': display_name,
                            'well_name': well_name
                        })
            
            # Check if we're in aggregated mode and use appropriate data
            if len(self.applied_wells) > 1 and self.df_aggregated_cache is not None:
                df_to_use = self.df_aggregated_cache
                well_name_to_use = self.df_aggregated_cache['Well_Name'].iloc[0]
            else:
                df_to_use = self.df_all
                well_name_to_use = self.current_well_name
            
            fig, results = run_arps_for_well_auto(
                df_to_use, well_name_to_use,
                outlier_threshold=self.current_analysis_params['outlier_threshold'],
                forecast_avg_points=self.current_analysis_params['forecast_avg_points'],
                manual_start_idx=self.current_analysis_params['manual_start_idx'],
                use_auto_detect=self.current_analysis_params['use_auto_detect'],
                start_method=self.current_analysis_params['start_method'],
                auto_start_params=self.current_analysis_params['auto_start_params'],
                filter_params=self.current_analysis_params['filter_params'],
                show_outliers=show_outliers,
                show_pre_decline=show_pre_decline,
                show_forecast=show_forecast,
                show_inliers=show_inliers,
                show_channel=show_channel,
                forecast_duration=self.current_analysis_params.get('forecast_duration', 60),
                fixed_qi=getattr(self, 'fixed_qi', None),
                reference_models=reference_models
            )
            if fig:
                self._display_figure(fig)
                
                # Redraw vertical line if in manual mode
                self.update_vertical_line_from_inputs()
                
                # Save updated plot options and references to session data
                self.save_current_settings_to_session()
                

        except Exception as e:
            print(f"Error updating plot options: {e}")
            import traceback
            traceback.print_exc()

    def open_range_config(self):
        from PyQt6.QtWidgets import QDialog, QDialogButtonBox
        
        config_dialog = QDialog(self)
        config_dialog.setWindowTitle("Configure Optimization Ranges")
        config_dialog.setMinimumSize(600, 500)
        layout = QVBoxLayout()
        config_dialog.setLayout(layout)
        
        # Single ranges widget
        ranges_widget = QWidget()
        ranges_layout = QGridLayout()
        ranges_widget.setLayout(ranges_layout)
        
        # Combine all ranges into a single dictionary
        all_ranges = {}
        all_vars = {}
        
        # Add filter ranges
        for param, (min_val, max_val) in self.filter_ranges.items():
            all_ranges[param] = (min_val, max_val)
        
        # Add auto start ranges
        for param, (min_val, max_val) in self.auto_start_ranges.items():
            all_ranges[param] = (min_val, max_val)
        
        # Create form for all ranges
        row = 0
        for param, (min_val, max_val) in all_ranges.items():
            ranges_layout.addWidget(QLabel(f"{param}:"), row, 0)
            ranges_layout.addWidget(QLabel("Min:"), row, 1)
            min_edit = QLineEdit(str(min_val))
            min_edit.setMaximumWidth(100)
            ranges_layout.addWidget(min_edit, row, 2)
            ranges_layout.addWidget(QLabel("Max:"), row, 3)
            max_edit = QLineEdit(str(max_val))
            max_edit.setMaximumWidth(100)
            ranges_layout.addWidget(max_edit, row, 4)
            all_vars[param] = (min_edit, max_edit)
            row += 1
        
        layout.addWidget(ranges_widget)
        
        # Buttons
        button_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.setProperty("class", "success")
        reset_btn = QPushButton("Reset to Default")
        reset_btn.setProperty("class", "secondary")
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setProperty("class", "secondary")
        
        def save_ranges():
            try:
                # Update filter ranges
                for param in self.filter_ranges.keys():
                    if param in all_vars:
                        min_edit, max_edit = all_vars[param]
                        self.filter_ranges[param] = (float(min_edit.text()), float(max_edit.text()))
                
                # Update auto start ranges
                for param in self.auto_start_ranges.keys():
                    if param in all_vars:
                        min_edit, max_edit = all_vars[param]
                        self.auto_start_ranges[param] = (float(min_edit.text()), float(max_edit.text()))
                
                QMessageBox.information(config_dialog, "Success", "Optimization ranges updated successfully!")
                config_dialog.accept()
            except Exception as e:
                QMessageBox.critical(config_dialog, "Error", f"Failed to update ranges: {str(e)}")
        
        def reset_ranges():
            # Reset filter ranges
            for param, (min_val, max_val) in DEFAULT_FILTER_RANGES.items():
                if param in all_vars:
                    all_vars[param][0].setText(str(min_val))
                    all_vars[param][1].setText(str(max_val))
            
            # Reset auto start ranges
            for param, (min_val, max_val) in DEFAULT_AUTO_START_RANGES.items():
                if param in all_vars:
                    all_vars[param][0].setText(str(min_val))
                    all_vars[param][1].setText(str(max_val))
        
        save_btn.clicked.connect(save_ranges)
        reset_btn.clicked.connect(reset_ranges)
        cancel_btn.clicked.connect(config_dialog.reject)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(reset_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        config_dialog.exec()

    def populate_well_tree(self):
        """Populate the tree widget with district -> field -> formation -> well hierarchy"""
        self.well_tree.clear()
        self._updating_tree = False  # Flag to prevent recursive updates
        
        for district in self.districts:
            district_item = QTreeWidgetItem([district])
            district_item.setFlags(district_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            district_item.setCheckState(0, Qt.CheckState.Unchecked)
            district_item.setData(0, Qt.ItemDataRole.UserRole, {'type': 'district', 'name': district})
            
            fields = self.fields_by_district.get(district, [])
            for field in fields:
                field_item = QTreeWidgetItem([field])
                field_item.setFlags(field_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                field_item.setCheckState(0, Qt.CheckState.Unchecked)
                field_item.setData(0, Qt.ItemDataRole.UserRole, {'type': 'field', 'name': field, 'district': district})
                
                formations = self.formations_by_field.get(f"{district}|{field}", [])
                for formation in formations:
                    formation_item = QTreeWidgetItem([formation])
                    formation_item.setFlags(formation_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    formation_item.setCheckState(0, Qt.CheckState.Unchecked)
                    formation_item.setData(0, Qt.ItemDataRole.UserRole, 
                                          {'type': 'formation', 'name': formation, 'district': district, 'field': field})
                    
                    wells = self.wells_by_formation.get(f"{district}|{field}|{formation}", [])
                    for well in wells:
                        well_item = QTreeWidgetItem([well])
                        well_item.setFlags(well_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                        well_item.setCheckState(0, Qt.CheckState.Unchecked)
                        well_item.setData(0, Qt.ItemDataRole.UserRole, 
                                         {'type': 'well', 'name': well, 'district': district, 
                                          'field': field, 'formation': formation})
                        formation_item.addChild(well_item)
                    
                    field_item.addChild(formation_item)
                district_item.addChild(field_item)
            
            self.well_tree.addTopLevelItem(district_item)
    
    def on_tree_item_changed(self, item, column):
        """Handle checkbox changes in the tree with parent/child relationships"""
        if self._updating_tree:
            return
        
        self._updating_tree = True
        
        # Update children when parent is checked/unchecked
        check_state = item.checkState(0)
        self._update_children(item, check_state)
        
        # Update parent state based on children
        parent = item.parent()
        if parent:
            self._update_parent_state(parent)
        
        self._updating_tree = False
    
    def _update_children(self, item, check_state):
        """Recursively update all children to match parent's check state"""
        for i in range(item.childCount()):
            child = item.child(i)
            child.setCheckState(0, check_state)
            self._update_children(child, check_state)
    
    def _update_parent_state(self, parent):
        """Update parent's check state based on children's states"""
        if not parent:
            return
        
        checked_count = 0
        unchecked_count = 0
        partial_count = 0
        total_children = parent.childCount()
        
        for i in range(total_children):
            child = parent.child(i)
            state = child.checkState(0)
            if state == Qt.CheckState.Checked:
                checked_count += 1
            elif state == Qt.CheckState.Unchecked:
                unchecked_count += 1
            elif state == Qt.CheckState.PartiallyChecked:
                partial_count += 1
        
        # Set parent state based on children's states
        if checked_count == total_children:
            # All children are fully checked
            parent.setCheckState(0, Qt.CheckState.Checked)
        elif unchecked_count == total_children:
            # All children are unchecked
            parent.setCheckState(0, Qt.CheckState.Unchecked)
        else:
            # Mixed state: some checked, some unchecked, or any partially checked
            parent.setCheckState(0, Qt.CheckState.PartiallyChecked)
        
        # Recursively update grandparent
        grandparent = parent.parent()
        if grandparent:
            self._update_parent_state(grandparent)
    
    def get_selected_wells(self):
        """Get list of all selected well names from the tree"""
        selected_wells = []
        
        def traverse_tree(item):
            data = item.data(0, Qt.ItemDataRole.UserRole)
            if data and data['type'] == 'well' and item.checkState(0) == Qt.CheckState.Checked:
                selected_wells.append(data['name'])
            
            for i in range(item.childCount()):
                traverse_tree(item.child(i))
        
        # Traverse all top-level items
        for i in range(self.well_tree.topLevelItemCount()):
            traverse_tree(self.well_tree.topLevelItem(i))
        
        return selected_wells
    
    def get_primary_well(self):
        """Get the first applied well (for backward compatibility with single-well operations)"""
        return self.applied_wells[0] if self.applied_wells else None
    
    def refresh_tree_display(self):
        """Refresh the tree widget display to show correct checkbox states using toggle trick"""
        # Set the updating flag to prevent triggering itemChanged events
        self._updating_tree = True
        
        # Toggle expansion state of root items to force visual refresh
        # This tricks Qt into redrawing the tree properly
        for i in range(self.well_tree.topLevelItemCount()):
            item = self.well_tree.topLevelItem(i)
            was_expanded = item.isExpanded()
            item.setExpanded(not was_expanded)
            item.setExpanded(was_expanded)
        
        # Force viewport update
        self.well_tree.viewport().update()
        
        # Reset the updating flag
        self._updating_tree = False
    
    def reset_parameters(self):
        """Reset all parameters to defaults except district/field/formation/well filters"""
        # Reset start method to Auto Select
        self.auto_select_radio.setChecked(True)
        
        # Reset general parameters to defaults
        self.threshold_edit.setText("2.0")
        self.threshold_opt_check.setChecked(True)
        self.forecast_avg_points_combo.setCurrentText("The Model")
        self.forecast_duration_edit.setText("60")
        
        # Reset auto start detection parameters to defaults
        self.smooth_window_edit.setText("13")
        self.smooth_window_opt_check.setChecked(True)
        self.min_segment_length_edit.setText("12")
        self.min_segment_length_opt_check.setChecked(True)
        self.change_sensitivity_edit.setText("0.3")
        self.change_sensitivity_opt_check.setChecked(True)
        self.stability_window_edit.setText("6")
        self.stability_window_opt_check.setChecked(True)
        self.min_decline_rate_edit.setText("0.15")
        self.min_decline_rate_opt_check.setChecked(True)
        
        # Reset plot options - show outliers, forecast, and pre-decline by default
        self.show_outliers_check.setChecked(True)
        self.show_pre_decline_check.setChecked(True)
        self.show_forecast_check.setChecked(True)
        self.show_inliers_check.setChecked(True)
        self.show_channel_check.setChecked(False)
        
        # Reset optimization checkbox
        self.optimize_check.setChecked(True)
        
        # Reset manual date inputs
        self.manual_year_edit.setText("2020")
        self.manual_month_edit.setText("1")
        self.qi_year_edit.setText("2020")
        self.qi_month_edit.setText("1")
        self.qi_value_edit.setText("")
        
        # Reset fixed_qi
        self.fixed_qi = None
        
        # Clear results text
        self.results_text.clear()
        
        # Clear session storage for applied wells (but keep applied_wells intact)
        for well_name in self.applied_wells:
            # Remove from session analyses if it exists
            if well_name in self.session_analyses:
                del self.session_analyses[well_name]
            # Remove from saved analyses tracking if it exists
            if well_name in self.saved_analyses:
                del self.saved_analyses[well_name]
        
        # Display raw data chart for currently applied wells (without clearing selection)
        if self.applied_wells:
            self.display_aggregated_raw_data(self.applied_wells)
        
        # Refresh tree display
        self.refresh_tree_display()
    
    def display_raw_data(self, well_name):
        """Display only raw production data points in black without any analysis"""
        df_well = self.df_all[self.df_all['Well_Name'] == well_name].copy()
        if df_well.empty:
            return
        
        df_well = df_well.sort_values('Prod_Date')
        df_well['Prod_Date'] = pd.to_datetime(df_well['Prod_Date'])
        df_well['days_in_month'] = df_well['Prod_Date'].apply(get_days_in_month)
        df_well['oil_prod_daily'] = df_well['M_Oil_Prod'] / df_well['days_in_month']
        
        # Create simple plot with only raw data
        fig, ax = plt.subplots(figsize=(12, 7))
        ax.plot(df_well['Prod_Date'], df_well['oil_prod_daily'], 'o', color='black', 
                label='Production Data', markersize=6)
        
        ax.set_xlabel('Date')
        ax.set_ylabel('Oil Production Rate (bbl/day)')
        ax.set_title(f'Well {well_name}')
        ax.legend()
        ax.grid(True, alpha=0.3)
        ax.set_ylim(bottom=0)
        fig.tight_layout()
        
        # Display the plot
        self._display_figure(fig)
        
        # Clear analysis params since this is just raw data
        self.current_well_name = None
        self.current_analysis_params = None

    def display_aggregated_raw_data(self, well_names):
        """Display aggregated raw production data for multiple wells without any analysis"""
        if not well_names:
            return
        
        # Create aggregated dataset
        df_aggregated = self.aggregate_well_data(well_names)
        if df_aggregated is None or df_aggregated.empty:
            return
        
        # Cache the aggregated data for chart click selection
        self.df_aggregated_cache = df_aggregated.copy()
        
        df_aggregated = df_aggregated.sort_values('Prod_Date')
        df_aggregated['Prod_Date'] = pd.to_datetime(df_aggregated['Prod_Date'])
        df_aggregated['days_in_month'] = df_aggregated['Prod_Date'].apply(get_days_in_month)
        df_aggregated['oil_prod_daily'] = df_aggregated['M_Oil_Prod'] / df_aggregated['days_in_month']
        
        # Create simple plot with only raw data
        fig, ax = plt.subplots(figsize=(12, 7))
        ax.plot(df_aggregated['Prod_Date'], df_aggregated['oil_prod_daily'], 'o', color='black', 
                label='Production Data', markersize=6)
        
        ax.set_xlabel('Date')
        ax.set_ylabel('Oil Production Rate (bbl/day)')
        
        # Create title based on selection coverage across formation/field/district
        title = self._build_selection_title(well_names)
        ax.set_title(title)
        
        ax.legend()
        ax.grid(True, alpha=0.3)
        ax.set_ylim(bottom=0)
        fig.tight_layout()
        
        # Display the plot
        self._display_figure(fig)
        
        # Clear analysis params since this is just raw data
        self.current_well_name = None
        self.current_analysis_params = None

    def aggregate_well_data(self, well_names):
        """Aggregate production data from multiple wells by summing monthly production"""
        if not well_names:
            return None
        
        # Get data for all selected wells
        df_combined = self.df_all[self.df_all['Well_Name'].isin(well_names)].copy()
        if df_combined.empty:
            return None
        
        # Convert date and sort
        df_combined['Prod_Date'] = pd.to_datetime(df_combined['Prod_Date'])
        
        # Group by date and sum the monthly production
        df_aggregated = df_combined.groupby('Prod_Date').agg({
            'M_Oil_Prod': 'sum',
            'District': 'first',  # Keep first district
            'Field': 'first',     # Keep first field
            'Alias_Formation': 'first'  # Keep first formation
        }).reset_index()
        
        # Create a synthetic well name for display
        if len(well_names) == 1:
            df_aggregated['Well_Name'] = well_names[0]
        else:
            df_aggregated['Well_Name'] = f"Aggregated ({len(well_names)} wells)"
        
        return df_aggregated

    def apply_selection(self):
        """Apply the current tree selection for analysis"""
        # Get selected wells
        selected_wells = self.get_selected_wells()
        
        if not selected_wells:
            QMessageBox.warning(self, "Warning", "Please select at least one well from the tree")
            return
        
        # Store the applied selection
        self.applied_wells = selected_wells
        
        # Update the status label
        if len(selected_wells) == 1:
            status_text = f"Applied: {selected_wells[0]}"
        else:
            status_text = f"Applied: {len(selected_wells)} wells"
        
        self.applied_selection_label.setText(status_text)
        self.applied_selection_label.setStyleSheet("""
            QLabel {
                color: #27ae60;
                font-weight: bold;
                font-size: 10px;
            }
        """)
        
        # Check if there's a matching analysis in the session
        matching_analysis = None
        for well_name, data in self.session_analyses.items():
            # Check if the applied_wells in the saved data matches current selection
            saved_wells = data.get('applied_wells', [well_name])
            if set(saved_wells) == set(selected_wells):
                matching_analysis = data
                break
        
        # If matching analysis found, restore it
        if matching_analysis:
            if self.apply_analysis_data(matching_analysis):
                # Re-run analysis to display the saved results
                QApplication.processEvents()
                self.run_analysis()
                
                # Restore references AFTER everything else is done
                self._restore_references_after_analysis()
                return
        
        # Otherwise, display the aggregated raw data on the chart
        self.display_aggregated_raw_data(selected_wells)
    
    def clear_selection(self):
        """Clear all checkboxes in the tree and reset the chart"""
        # Temporarily disconnect signal to avoid triggering events
        self.well_tree.itemChanged.disconnect(self.on_tree_item_changed)
        
        # Uncheck all items in the tree
        def uncheck_all(item):
            item.setCheckState(0, Qt.CheckState.Unchecked)
            for i in range(item.childCount()):
                uncheck_all(item.child(i))
        
        for i in range(self.well_tree.topLevelItemCount()):
            uncheck_all(self.well_tree.topLevelItem(i))
        
        # Reconnect signal
        self.well_tree.itemChanged.connect(self.on_tree_item_changed)
        
        # Clear applied selection
        self.applied_wells = []
        self.df_aggregated_cache = None
        self.applied_selection_label.setText("No selection applied")
        self.applied_selection_label.setStyleSheet("""
            QLabel {
                color: #7f8c8d;
                font-style: italic;
                font-size: 10px;
            }
        """)
        
        # Clear the chart
        self._cleanup_canvas_and_toolbar()
        self.figure = None
        self.current_well_name = None
        self.current_analysis_params = None

    def run_analysis(self):
        """Run analysis on applied selection (single or aggregated wells)"""
        if not self.applied_wells:
            QMessageBox.warning(self, "Warning", "Please apply a selection first using the 'Apply Selection' button")
            return
        
        # Create aggregated dataset
        print(f"DEBUG: run_analysis called with applied_wells={self.applied_wells}")
        df_aggregated = self.aggregate_well_data(self.applied_wells)
        print(f"DEBUG: aggregate_well_data returned: {df_aggregated is not None and not df_aggregated.empty}")
        if df_aggregated is None or df_aggregated.empty:
            print(f"DEBUG: Applied wells: {self.applied_wells}")
            print(f"DEBUG: Available wells in df_all: {self.df_all['Well_Name'].unique()[:10]}")
            QMessageBox.warning(self, "Warning", "No data found for applied wells")
            return
        
        # Cache the aggregated data for chart click selection
        self.df_aggregated_cache = df_aggregated.copy()
        
        # Temporarily replace df_all with aggregated data for analysis
        original_df = self.df_all
        self.df_all = df_aggregated
        
        try:
            # Run the analysis with the aggregated data
            well_name_for_analysis = df_aggregated['Well_Name'].iloc[0]
            print(f"DEBUG: Running analysis with well_name={well_name_for_analysis}")
            print(f"DEBUG: df_all now has wells: {self.df_all['Well_Name'].unique()}")
            self.run_analysis_internal(well_name_for_analysis)
        finally:
            # Restore original dataframe
            self.df_all = original_df

    def run_analysis_internal(self, well_name):
        print(f"DEBUG: run_analysis_internal called with well_name='{well_name}'")
        if not well_name:
            QMessageBox.warning(self, "Warning", "Please select a well")
            return

        # Get start method
        if self.auto_select_radio.isChecked():
            start_method = "Auto Select"
        elif self.manual_start_radio.isChecked():
            start_method = "Manual Start Date"
        else:
            start_method = "Manual Start Date and Initial Rate (Qi)"
            
        manual_start_idx = None
        method_label = start_method
        df_well_temp = self.df_all[self.df_all['Well_Name'] == well_name].copy()
        df_well_temp['Prod_Date'] = pd.to_datetime(df_well_temp['Prod_Date'])
        df_well_sorted = df_well_temp.sort_values('Prod_Date')

        if start_method != "Manual Start Date and Initial Rate (Qi)":
            self.fixed_qi = None

        if start_method in ["Manual Start Date", "Manual Start Date and Initial Rate (Qi)"]:
            try:
                if start_method == "Manual Start Date":
                    year = int(self.manual_year_edit.text())
                    month = int(self.manual_month_edit.text())
                else:
                    year = int(self.qi_year_edit.text())
                    month = int(self.qi_month_edit.text())
                    # Read Qi value from entry field
                    qi_str = self.qi_value_edit.text().strip()
                    if qi_str:
                        try:
                            self.fixed_qi = float(qi_str)
                        except ValueError:
                            QMessageBox.critical(self, "Error", f"Invalid Initial Rate (Qi) value: {qi_str}")
                            return
                    else:
                        QMessageBox.warning(self, "Warning", "Please enter an Initial Rate (Qi) value or click on the chart to select one.")
                        return
                manual_date = pd.to_datetime(f"{year}-{month:02d}-01")
                manual_start_idx = np.searchsorted(df_well_sorted['Prod_Date'], manual_date)
                if manual_start_idx >= len(df_well_sorted):
                    QMessageBox.warning(self, "Warning", "Manual date is after last production date. Using last index.")
                    manual_start_idx = len(df_well_sorted) - 1
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Invalid manual date: {str(e)}")
                return

        outlier_threshold = float(self.threshold_edit.text())
        
        # Convert forecast_avg_points combo selection to integer value
        forecast_avg_text = self.forecast_avg_points_combo.currentText()
        if forecast_avg_text == "The Model":
            forecast_avg_points = 0
        elif forecast_avg_text == "The Last Rate":
            forecast_avg_points = 1
        else:
            forecast_avg_points = int(forecast_avg_text)
        
        forecast_duration = int(self.forecast_duration_edit.text())

        filter_params = {'threshold': float(self.threshold_edit.text())}
        auto_start_params = {
            'smooth_window': int(self.smooth_window_edit.text()),
            'min_segment_length': int(self.min_segment_length_edit.text()),
            'change_sensitivity': float(self.change_sensitivity_edit.text()),
            'stability_window': int(self.stability_window_edit.text()),
            'min_decline_rate': float(self.min_decline_rate_edit.text())
        }

        show_outliers = self.show_outliers_check.isChecked()
        show_pre_decline = self.show_pre_decline_check.isChecked()
        show_forecast = self.show_forecast_check.isChecked()
        show_inliers = self.show_inliers_check.isChecked()
        show_channel = self.show_channel_check.isChecked()
        optimize = self.optimize_check.isChecked()
        
        # Get reference model parameters
        reference_models = []
        selected_items = self.reference_list.selectedItems()
        for item in selected_items:
            ref_well_name = item.data(Qt.ItemDataRole.UserRole)
            if ref_well_name in self.session_analyses:
                data = self.session_analyses[ref_well_name]
                if 'qi' in data and 'Di' in data and 'b' in data:
                    qi = data['qi']
                    Di = data['Di']
                    b = data['b']
                    # Create reference model dict with well name for display
                    hierarchy_parts = ref_well_name.split('\\')
                    if len(hierarchy_parts) >= 3:
                        display_name = f"{hierarchy_parts[-3]}\\{hierarchy_parts[-2]}\\{hierarchy_parts[-1]}"
                    else:
                        display_name = ref_well_name
                    
                    reference_models.append({
                        'Di': Di, 
                        'b': b, 
                        'name': display_name,
                        'well_name': ref_well_name
                    })

        pbounds = {}
        param_to_opt_check = {
            'threshold': 'threshold_opt_check',
            'smooth_window': 'smooth_window_opt_check',
            'min_segment_length': 'min_segment_length_opt_check',
            'change_sensitivity': 'change_sensitivity_opt_check',
            'stability_window': 'stability_window_opt_check',
            'min_decline_rate': 'min_decline_rate_opt_check'
        }
        relevant_ranges = {**self.filter_ranges, **self.auto_start_ranges}
        for param in relevant_ranges:
            opt_check = param_to_opt_check.get(param)
            if opt_check and hasattr(self, opt_check) and getattr(self, opt_check).isChecked():
                pbounds[param] = relevant_ranges[param]

        if optimize:
            fixed_params = {**filter_params, **auto_start_params}
            if start_method == "Auto Select":
                chosen_idx = find_optimal_start_date(df_well_sorted, **auto_start_params)
                method_label = "Auto-Detected Optimal Start"
                n_iter = 15
            elif start_method in ["Manual Start Date", "Manual Start Date and Initial Rate (Qi)"]:
                chosen_idx = manual_start_idx
                n_iter = 10
            else:
                chosen_idx = 0
                n_iter = 0

            if pbounds:
                def objective_function(**params):
                    for param in int_keys:
                        if param in params:
                            params[param] = int(round(params[param]))
                    sampled_filter = {k: params.get(k, fixed_params[k]) for k in self.filter_ranges}
                    sampled_auto_start = {k: params.get(k, fixed_params[k]) for k in self.auto_start_ranges}

                    if start_method == "Auto Select":
                        start_idx = find_optimal_start_date(df_well_sorted, **sampled_auto_start)
                    elif start_method in ["Manual Start Date", "Manual Start Date and Initial Rate (Qi)"]:
                        start_idx = manual_start_idx
                    else:
                        start_idx = 0

                    metrics = compute_metrics(
                        self.df_all, well_name, manual_start_idx=start_idx,
                        filter_params=sampled_filter,
                        auto_start_params=sampled_auto_start,
                        ignore_gaps=(start_method in ["Manual Start Date", "Manual Start Date and Initial Rate (Qi)"]),
                        fixed_qi=self.fixed_qi if start_method == "Manual Start Date and Initial Rate (Qi)" else None
                    )

                    if 'popt' in metrics and metrics['popt'] is not None:
                        Di = metrics['popt'][1]
                        if Di < 0.001:
                            metrics['score'] *= 0.01
                        elif Di < 0.01:
                            metrics['score'] *= 0.3
                        elif Di < 0.05:
                            metrics['score'] *= 0.7

                    # Ensure we never return infinity values for Bayesian optimization
                    score = metrics['score']
                    if not np.isfinite(score):
                        score = -1e10
                    return score

                try:
                    optimizer = BayesianOptimization(
                        f=objective_function,
                        pbounds=pbounds,
                        random_state=1,
                        verbose=0
                    )
                    optimizer.maximize(init_points=5, n_iter=n_iter)
                    best_params = optimizer.max['params']
                    best_score = optimizer.max['target']
                    for param in int_keys:
                        if param in best_params:
                            best_params[param] = int(round(best_params[param]))
                    optimized_params = []
                    for param in best_params:
                        opt_check = param_to_opt_check.get(param)
                        if opt_check and hasattr(self, opt_check) and getattr(self, opt_check).isChecked():
                            edit_name = f"{param}_edit"
                            if hasattr(self, edit_name):
                                value = best_params[param]
                                if param in int_keys:
                                    getattr(self, edit_name).setText(str(int(value)))
                                else:
                                    getattr(self, edit_name).setText(str(round(value, 2)))
                                optimized_params.append(param)
                            if param in self.filter_ranges:
                                filter_params[param] = best_params[param]
                            elif param in self.auto_start_ranges:
                                auto_start_params[param] = best_params[param]
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Optimization failed: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    return

                fig, results = run_arps_for_well_auto(
                    self.df_all, well_name, outlier_threshold=outlier_threshold,
                    forecast_avg_points=forecast_avg_points, manual_start_idx=chosen_idx,
                    use_auto_detect=False, start_method=method_label,
                    auto_start_params=auto_start_params,
                    filter_params=filter_params,
                    show_outliers=show_outliers, show_pre_decline=show_pre_decline,
                    show_forecast=show_forecast, show_inliers=show_inliers,
                    show_channel=show_channel, forecast_duration=forecast_duration,
                    fixed_qi=self.fixed_qi if start_method == "Manual Start Date and Initial Rate (Qi)" else None,
                    reference_models=reference_models
                )
            else:
                fig, results = run_arps_for_well_auto(
                    self.df_all, well_name, outlier_threshold=outlier_threshold,
                    forecast_avg_points=forecast_avg_points, manual_start_idx=chosen_idx,
                    use_auto_detect=False, start_method=method_label,
                    auto_start_params=auto_start_params,
                    filter_params=filter_params,
                    show_outliers=show_outliers, show_pre_decline=show_pre_decline,
                    show_forecast=show_forecast, show_inliers=show_inliers,
                    show_channel=show_channel, forecast_duration=forecast_duration,
                    fixed_qi=self.fixed_qi if start_method == "Manual Start Date and Initial Rate (Qi)" else None,
                    reference_models=reference_models
                )
        else:
            if start_method == "Auto Select":
                chosen_idx = find_optimal_start_date(df_well_sorted, **auto_start_params)
                method_label = "Auto-Detected Optimal Start"
                fig, results = run_arps_for_well_auto(
                    self.df_all, well_name, outlier_threshold=outlier_threshold,
                    forecast_avg_points=forecast_avg_points, manual_start_idx=chosen_idx,
                    use_auto_detect=False, start_method=method_label,
                    auto_start_params=auto_start_params,
                    filter_params=filter_params,
                    show_outliers=show_outliers, show_pre_decline=show_pre_decline,
                    show_forecast=show_forecast, show_inliers=show_inliers,
                    show_channel=show_channel, forecast_duration=forecast_duration,
                    fixed_qi=None,
                    reference_models=reference_models
                )
            elif start_method in ["Manual Start Date", "Manual Start Date and Initial Rate (Qi)"]:
                fig, results = run_arps_for_well_auto(
                    self.df_all, well_name, outlier_threshold=outlier_threshold,
                    forecast_avg_points=forecast_avg_points, manual_start_idx=manual_start_idx,
                    use_auto_detect=False, start_method=start_method,
                    auto_start_params=auto_start_params,
                    filter_params=filter_params,
                    show_outliers=show_outliers, show_pre_decline=show_pre_decline,
                    show_forecast=show_forecast, show_inliers=show_inliers,
                    show_channel=show_channel, forecast_duration=forecast_duration,
                    fixed_qi=self.fixed_qi if start_method == "Manual Start Date and Initial Rate (Qi)" else None,
                    reference_models=reference_models
                )

        # Override title to reflect grouping-aware selection
        if fig is not None and getattr(fig, 'axes', None):
            fig.axes[0].set_title(self._build_selection_title(self.applied_wells))

        self.results_text.clear()
        self.results_text.setText(results)

        if fig:
            self._display_figure(fig)
            
            self.current_well_name = well_name
            self.current_analysis_params = {
                'outlier_threshold': outlier_threshold,
                'forecast_avg_points': forecast_avg_points,
                'manual_start_idx': chosen_idx if 'chosen_idx' in locals() else manual_start_idx,
                'use_auto_detect': False,
                'start_method': method_label if 'method_label' in locals() else start_method,
                'auto_start_params': auto_start_params,
                'filter_params': filter_params,
                'forecast_duration': forecast_duration
            }
            
            # Redraw vertical line if in manual mode
            self.update_vertical_line_from_inputs()
            
            # Store analysis in session
            analysis_data = self.create_analysis_data(well_name)
            if analysis_data:
                # Extract qi, Di, b from results_text and add to session data
                results_text = self.results_text.toPlainText()
                qi_val = None
                di_val = None
                b_val = None
                
                for line in results_text.split('\n'):
                    if 'qi =' in line and 'Di =' in line and 'b =' in line:
                        try:
                            parts = line.split(',')
                            qi_val = float(parts[0].split('=')[1].strip())
                            di_val = float(parts[1].split('=')[1].strip())
                            b_val = float(parts[2].split('=')[1].strip())
                        except:
                            pass
                
                if qi_val is not None:
                    analysis_data['qi'] = qi_val
                    analysis_data['Di'] = di_val
                    analysis_data['b'] = b_val
                
                # Store the decline start date
                if self.current_analysis_params and well_name:
                    start_idx = self.current_analysis_params.get('manual_start_idx', 0)
                    if len(self.applied_wells) > 1 and self.df_aggregated_cache is not None:
                        df_for_date = self.df_aggregated_cache
                    else:
                        df_for_date = self.df_all
                    
                    df_temp = df_for_date[df_for_date['Well_Name'] == well_name].copy()
                    if not df_temp.empty:
                        df_temp = df_temp.sort_values('Prod_Date')
                        if start_idx > 0 and start_idx < len(df_temp):
                            df_temp = df_temp.iloc[start_idx:]
                        if len(df_temp) > 0:
                            analysis_data['decline_start_date'] = df_temp.iloc[0]['Prod_Date'].isoformat()
                
                self.session_analyses[well_name] = analysis_data
                self.saved_analyses[well_name] = False  # Mark as unsaved
                
                # Update reference list
                self.update_reference_list()
        else:
            QMessageBox.critical(self, "Error", results)
    
    def create_export_sheet(self, wb, sheet_name, analysis_dict, export_history, export_forecast, forecast_duration, is_monthly, use_auto_fsd=True, fixed_fsd_date=None):
        """Helper method to create an export sheet for analyses"""
        # Group analyses by formation
        formation_groups = {}
        for analysis_name, analysis_info in analysis_dict.items():
            formation = analysis_info.get('formation', 'Unknown')
            if formation not in formation_groups:
                formation_groups[formation] = []
            formation_groups[formation].append((analysis_name, analysis_info))
        
        # Collect all dates (history + forecast) with frequency filtering
        all_dates_list = []
        
        # Add history dates if requested
        if export_history:
            for analysis_name, analysis_info in analysis_dict.items():
                dates = analysis_info['data']['Prod_Date'].tolist()
                if is_monthly:
                    # Use all monthly dates
                    all_dates_list.extend(dates)
                else:
                    # For yearly, collect all months (we'll average by year later)
                    all_dates_list.extend(dates)
        
        # Add forecast dates if requested
        if export_forecast:
            for analysis_name, analysis_info in analysis_dict.items():
                analysis_params = analysis_info['analysis'].get('analysis_params')
                if analysis_params:
                    # Get model parameters from analysis
                    results_text = analysis_info['analysis'].get('results_text', '')
                    qi = None
                    Di = None
                    b = None
                    
                    for line in results_text.split('\n'):
                        if 'qi =' in line and 'Di =' in line and 'b =' in line:
                            try:
                                parts = line.split(',')
                                qi = float(parts[0].split('=')[1].strip())
                                Di = float(parts[1].split('=')[1].strip())
                                b = float(parts[2].split('=')[1].strip())
                            except:
                                pass
                    
                    if qi is not None and Di is not None and b is not None:
                        # Generate forecast dates
                        max_date = analysis_info['data']['Prod_Date'].max()
                        
                        # Determine forecast start date (fsd)
                        if use_auto_fsd:
                            # Use auto: month after last history point for each analysis
                            fsd = max_date + pd.DateOffset(months=1)
                        else:
                            # Use fixed fsd for all analyses
                            fsd = fixed_fsd_date
                        
                        forecast_dates = pd.date_range(
                            start=fsd,
                            periods=forecast_duration,
                            freq='MS'
                        )
                        
                        # Apply frequency filtering
                        if is_monthly:
                            all_dates_list.extend(forecast_dates.tolist())
                        else:
                            # For yearly, collect all months (we'll average by year later)
                            all_dates_list.extend(forecast_dates.tolist())
        
        # For yearly frequency, group dates by year and create representative dates
        if not is_monthly:
            # Group by year
            dates_by_year = {}
            for date in all_dates_list:
                year = date.year
                if year not in dates_by_year:
                    dates_by_year[year] = []
                dates_by_year[year].append(date)
            
            # Create one date per year (use January 1st as representative)
            all_dates = sorted([pd.Timestamp(year, 1, 1) for year in dates_by_year.keys()])
        else:
            all_dates = sorted(set(all_dates_list))
        
        if not all_dates:
            return  # Skip if no dates
        
        # Create sheet for this field
        ws = wb.create_sheet(title=sheet_name)
        
        # Write header
        row = 1
        col = 1
        ws.cell(row, col, 'Date')
        
        # Write analysis names as headers grouped by formation
        for formation in sorted(formation_groups.keys()):
            analyses_in_formation = formation_groups[formation]
            for analysis_name, _ in analyses_in_formation:
                col += 1
                ws.cell(row, col, analysis_name)
        
        # Write data
        for date in all_dates:
            col = 1
            ws.cell(row + 1, col, date.strftime('%Y-%m-%d'))
            
            # Determine if this is forecast or history
            max_history_date = None
            if export_history and analysis_dict:
                try:
                    max_history_date = max(analysis_info['data']['Prod_Date'].max() for analysis_info in analysis_dict.values())
                except:
                    pass
            
            is_forecast = False
            if export_forecast and max_history_date is not None:
                is_forecast = date > max_history_date
            elif export_forecast and not export_history:
                # If only exporting forecast, check if date is after the latest in data
                try:
                    max_date_in_data = max(analysis_info['data']['Prod_Date'].max() for analysis_info in analysis_dict.values())
                    is_forecast = date > max_date_in_data
                except:
                    pass
            
            # Fill in data for each analysis
            for formation in sorted(formation_groups.keys()):
                analyses_in_formation = formation_groups[formation]
                for analysis_name, analysis_info in analyses_in_formation:
                    col += 1
                    
                    if is_forecast and export_forecast:
                        # Use the EXACT qi, Di, b values from the analysis
                        qi = analysis_info['analysis'].get('qi')
                        Di = analysis_info['analysis'].get('Di')
                        b = analysis_info['analysis'].get('b')
                        
                        # If not stored, try to extract from results_text
                        if qi is None or Di is None or b is None:
                            results_text = analysis_info['analysis'].get('results_text', '')
                            for line in results_text.split('\n'):
                                if 'qi =' in line and 'Di =' in line and 'b =' in line:
                                    try:
                                        parts = line.split(',')
                                        qi = float(parts[0].split('=')[1].strip())
                                        Di = float(parts[1].split('=')[1].strip())
                                        b = float(parts[2].split('=')[1].strip())
                                    except:
                                        pass
                        
                        if qi is not None and Di is not None and b is not None:
                            # Extract forecast_avg_points for this specific analysis
                            analysis_data = analysis_info['analysis']
                            forecast_avg_text = analysis_data.get('forecast_avg_points', 'The Model')
                            if forecast_avg_text == "The Model":
                                analysis_forecast_avg_points = 0
                            elif forecast_avg_text == "The Last Rate":
                                analysis_forecast_avg_points = 1
                            else:
                                try:
                                    analysis_forecast_avg_points = int(forecast_avg_text)
                                except:
                                    analysis_forecast_avg_points = 0
                            
                            # Get the EXACT decline start date from analysis
                            decline_start = analysis_data.get('decline_start_date')
                            if decline_start:
                                decline_start = pd.to_datetime(decline_start)
                            else:
                                # Fallback: use stored start_date or min date
                                decline_start = analysis_info.get('start_date', analysis_info['data']['Prod_Date'].min())
                                if isinstance(decline_start, str):
                                    decline_start = pd.to_datetime(decline_start)
                            
                            last_date = analysis_info['data']['Prod_Date'].max()
                            
                            # Calculate last_t: time from EXACT decline start to last historical data point
                            last_t = (last_date.year - decline_start.year) * 12 + (last_date.month - decline_start.month)
                            
                            # Calculate months ahead from last date
                            if not use_auto_fsd and fixed_fsd_date is not None:
                                # Using fixed fsd - treat any gap as a single timestep
                                gap_months = ((fixed_fsd_date.year - last_date.year) * 12 + 
                                            (fixed_fsd_date.month - last_date.month))
                                if gap_months > 1:
                                    # There's a gap - treat it as 1 timestep
                                    months_from_fsd = ((date.year - fixed_fsd_date.year) * 12 + 
                                                      (date.month - fixed_fsd_date.month))
                                    months_ahead = 1 + max(0, months_from_fsd)
                                else:
                                    # No gap or small gap - normal calculation
                                    months_ahead = (date.year - last_date.year) * 12 + (date.month - last_date.month)
                            else:
                                # Using auto fsd - normal calculation
                                months_ahead = (date.year - last_date.year) * 12 + (date.month - last_date.month)
                            
                            months_ahead = max(0, months_ahead)  # Ensure t continues from history
                            
                            # For forecast: t = last_t + months_ahead
                            t = last_t + months_ahead
                            
                            forecast_value = arps_hyperbolic(t, qi, Di, b)
                            
                            # Apply forecast adjustments if needed (avg_points)
                            if analysis_forecast_avg_points > 1:
                                # Get last n points from data for adjustment
                                data_df = analysis_info['data'].sort_values('Prod_Date')
                                n_points = min(analysis_forecast_avg_points, len(data_df))
                                if n_points > 0:
                                    last_rates = data_df['oil_prod_daily'].tail(n_points).values
                                    actual_avg = np.mean(last_rates)
                                    fitted_rate = arps_hyperbolic(last_t, qi, Di, b)
                                    adjustment = actual_avg - fitted_rate
                                    forecast_value = max(0, forecast_value + adjustment)
                            elif analysis_forecast_avg_points == 1:
                                # Use last historical rate
                                data_df = analysis_info['data'].sort_values('Prod_Date')
                                if len(data_df) > 0:
                                    last_rate = data_df['oil_prod_daily'].iloc[-1]
                                    fitted_rate = arps_hyperbolic(last_t, qi, Di, b)
                                    adjustment = last_rate - fitted_rate
                                    forecast_value = max(0, forecast_value + adjustment)
                            
                            # For yearly frequency, calculate average for all months in that year
                            if not is_monthly:
                                # Calculate value for all future months in this year and average
                                year_values = []
                                for month in range(1, 13):
                                    month_date = pd.Timestamp(date.year, month, 1)
                                    
                                    # Only forecast months that are after last_date
                                    if month_date <= last_date:
                                        continue
                                    
                                    # Calculate months ahead for this month (with gap handling for fixed fsd)
                                    if not use_auto_fsd and fixed_fsd_date is not None:
                                        gap_months = ((fixed_fsd_date.year - last_date.year) * 12 + 
                                                    (fixed_fsd_date.month - last_date.month))
                                        if gap_months > 1:
                                            # There's a gap - treat it as 1 timestep
                                            months_from_fsd = ((month_date.year - fixed_fsd_date.year) * 12 + 
                                                              (month_date.month - fixed_fsd_date.month))
                                            months_ahead_month = 1 + max(0, months_from_fsd)
                                        else:
                                            # No gap or small gap - normal calculation
                                            months_ahead_month = (month_date.year - last_date.year) * 12 + (month_date.month - last_date.month)
                                    else:
                                        # Using auto fsd - normal calculation
                                        months_ahead_month = (month_date.year - last_date.year) * 12 + (month_date.month - last_date.month)
                                    
                                    months_ahead_month = max(0, months_ahead_month)  # Ensure t continues from history
                                    t_month = last_t + months_ahead_month
                                    forecast_value_month = arps_hyperbolic(t_month, qi, Di, b)
                                    
                                    # Apply adjustments if needed
                                    if analysis_forecast_avg_points > 1:
                                        data_df = analysis_info['data'].sort_values('Prod_Date')
                                        n_points = min(analysis_forecast_avg_points, len(data_df))
                                        if n_points > 0:
                                            last_rates = data_df['oil_prod_daily'].tail(n_points).values
                                            actual_avg = np.mean(last_rates)
                                            fitted_rate = arps_hyperbolic(last_t, qi, Di, b)
                                            adjustment = actual_avg - fitted_rate
                                            forecast_value_month = max(0, forecast_value_month + adjustment)
                                    elif analysis_forecast_avg_points == 1:
                                        data_df = analysis_info['data'].sort_values('Prod_Date')
                                        if len(data_df) > 0:
                                            last_rate = data_df['oil_prod_daily'].iloc[-1]
                                            fitted_rate = arps_hyperbolic(last_t, qi, Di, b)
                                            adjustment = last_rate - fitted_rate
                                            forecast_value_month = max(0, forecast_value_month + adjustment)
                                    
                                    year_values.append(forecast_value_month)
                                
                                if year_values:
                                    forecast_value = np.mean(year_values)
                            
                            ws.cell(row + 1, col, forecast_value)
                    elif not is_forecast and export_history:
                        # Use actual production data
                        analysis_data = analysis_info['data']
                        
                        if is_monthly:
                            # For monthly, match exact date
                            matching_row = analysis_data[analysis_data['Prod_Date'] == date]
                            if not matching_row.empty:
                                ws.cell(row + 1, col, matching_row.iloc[0]['oil_prod_daily'])
                        else:
                            # For yearly, average all months in that year
                            year_data = analysis_data[analysis_data['Prod_Date'].dt.year == date.year]
                            if not year_data.empty:
                                avg_value = year_data['oil_prod_daily'].mean()
                                ws.cell(row + 1, col, avg_value)
            
            row += 1
        
        # Auto-fit columns based on content
        for col_idx in range(1, col + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            # Check header
            header = ws.cell(1, col_idx).value
            if header:
                max_length = max(max_length, len(str(header)))
            # Check data rows
            for row_idx in range(2, row):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width (min 10, max 50, with some padding)
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
    
    def calculate_decline_percentage_model(self, t_values, qi, Di, b, years_back):
        """Calculate average decline percentage for the last N years using fitted model"""
        if len(t_values) == 0:
            return ''
        
        # Get the time period
        max_t = t_values[-1]
        min_t = max(0, max_t - years_back * 12)  # years_back in months
        
        # Calculate model values at start and end of period
        if min_t >= max_t:
            return ''
        
        start_rate = arps_hyperbolic(min_t, qi, Di, b)
        end_rate = arps_hyperbolic(max_t, qi, Di, b)
        
        if start_rate == 0 or not np.isfinite(start_rate) or not np.isfinite(end_rate):
            return ''
        
        decline_pct = ((start_rate - end_rate) / start_rate) * 100
        return f"{decline_pct:.2f}"
    
    def update_analysis_list(self, index):
        """Update the analysis list when exports tab is opened"""
        # Only update if we're on the exports tab (index 2: Analysis, Setup, Exports)
        if index == 2:  # Exports tab
            self.update_analysis_list_for_export()
    
    def update_analysis_list_for_export(self):
        """Update the analysis list in the export tab"""
        self.analysis_list.clear()
        
        # Populate with available analyses
        for well_name, data in sorted(self.session_analyses.items()):
            applied_wells = data.get('applied_wells', [well_name])
            hierarchy_name = self._build_selection_title(applied_wells)
            
            # Add item to list
            self.analysis_list.addItem(hierarchy_name)
            
            # Get the item we just added and configure it
            item = self.analysis_list.item(self.analysis_list.count() - 1)
            # Store well_name as item data
            item.setData(Qt.ItemDataRole.UserRole, well_name)
            
            # Select all by default
            item.setSelected(True)
    
    def select_all_analyses(self):
        """Select all analyses in the list"""
        for i in range(self.analysis_list.count()):
            self.analysis_list.item(i).setSelected(True)
    
    def deselect_all_analyses(self):
        """Deselect all analyses in the list"""
        for i in range(self.analysis_list.count()):
            self.analysis_list.item(i).setSelected(False)
    
    def export_to_excel(self):
        """Export all analyses to Excel with sheets organized by Field"""
        if not self.session_analyses:
            QMessageBox.warning(self, "Warning", "No analyses to export. Please run analyses first.")
            return
        
        # Get selected analyses
        selected_items = self.analysis_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", "Please select at least one analysis to export.")
            return
        
        selected_well_names = [item.data(Qt.ItemDataRole.UserRole) for item in selected_items]
        selected_analyses = {k: v for k, v in self.session_analyses.items() if k in selected_well_names}
        
        # Get export options
        export_history = self.export_history_check.isChecked()
        export_forecast = self.export_forecast_check.isChecked()
        
        if not export_history and not export_forecast:
            QMessageBox.warning(self, "Warning", "Please select at least one export option (History or Forecast).")
            return
        
        # Get forecast parameters from Setup tab
        forecast_duration = int(self.forecast_duration_edit.text()) if self.forecast_duration_edit.text() else 60
        
        # Get forecast start date (fsd) settings from Export tab
        use_auto_fsd = self.fsd_auto_radio.isChecked()
        fixed_fsd_date = None
        if not use_auto_fsd:
            fsd_text = self.fsd_date_edit.text()
            if fsd_text:
                try:
                    fixed_fsd_date = pd.to_datetime(fsd_text)
                except:
                    QMessageBox.warning(self, "Warning", "Invalid date format. Using auto fsd for all analyses.")
                    use_auto_fsd = True
        
        is_monthly = self.monthly_radio.isChecked()
        include_summary = self.include_summary_check.isChecked()
        
        # Open file dialog
        default_filename = f"Decline_Curve_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*.*)"
        )
        
        if not filepath:
            return  # User cancelled
        
        # Ensure .xlsx extension
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'
        
        try:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Organize analyses by field and separate aggregated analyses
            field_data = {}  # {field_name: {analysis_name: {data, analysis}}}
            aggregated_data = {}  # {analysis_name: {data, analysis}}
            
            for well_name, analysis_data in selected_analyses.items():
                applied_wells = analysis_data.get('applied_wells', [well_name])
                
                # For aggregated wells, use the aggregated data
                if len(applied_wells) > 1:
                    # This is an aggregated analysis - put in aggregated_data
                    df_aggregated = self.aggregate_well_data(applied_wells)
                    
                    if not df_aggregated.empty:
                        df_aggregated = df_aggregated.sort_values('Prod_Date')
                        df_aggregated['Prod_Date'] = pd.to_datetime(df_aggregated['Prod_Date'])
                        df_aggregated['days_in_month'] = df_aggregated['Prod_Date'].apply(get_days_in_month)
                        df_aggregated['oil_prod_daily'] = df_aggregated['M_Oil_Prod'] / df_aggregated['days_in_month']
                        
                        # Use the analysis name (which shows the aggregated wells)
                        analysis_display_name = self._build_selection_title(applied_wells)
                        aggregated_data[analysis_display_name] = {
                            'data': df_aggregated[['Prod_Date', 'oil_prod_daily']],
                            'analysis': analysis_data,
                            'applied_wells': applied_wells,
                            'start_date': df_aggregated['Prod_Date'].min()
                        }
                else:
                    # Single well analysis - put in field_data
                    df_well = self.df_master[self.df_master['Well_Name'] == applied_wells[0]].copy()
                    
                    if not df_well.empty:
                        df_well = df_well.sort_values('Prod_Date')
                        df_well['Prod_Date'] = pd.to_datetime(df_well['Prod_Date'])
                        df_well['days_in_month'] = df_well['Prod_Date'].apply(get_days_in_month)
                        df_well['oil_prod_daily'] = df_well['M_Oil_Prod'] / df_well['days_in_month']
                        
                        # Get the actual field/formation for this well from the data
                        district = df_well.iloc[0]['District'] if 'District' in df_well.columns else ''
                        field = df_well.iloc[0]['Field'] if 'Field' in df_well.columns else ''
                        formation = df_well.iloc[0]['Alias_Formation'] if 'Alias_Formation' in df_well.columns else ''
                        
                        if field not in field_data:
                            field_data[field] = {}
                        
                        field_data[field][applied_wells[0]] = {
                            'data': df_well[['Prod_Date', 'oil_prod_daily']],
                            'analysis': analysis_data,
                            'formation': formation,
                            'district': district,
                            'start_date': df_well['Prod_Date'].min()
                        }
            
            # Create sheet for aggregated analyses first (if any)
            if aggregated_data:
                # Create "Aggregated" sheet
                self.create_export_sheet(wb, "Aggregated", aggregated_data, export_history, export_forecast, forecast_duration, is_monthly, use_auto_fsd, fixed_fsd_date)
            
            # Create sheets for each field with single-well analyses only
            for field_name, analysis_dict in sorted(field_data.items()):
                self.create_export_sheet(wb, field_name, analysis_dict, export_history, export_forecast, forecast_duration, is_monthly, use_auto_fsd, fixed_fsd_date)
            
            # Create summary sheet if requested
            if include_summary:
                summary_ws = wb.create_sheet(title="Summary", index=0)
                
                # Summary header with decline percentages
                headers = ['Well Name', 'District', 'Field', 'Formation', 
                          'Decline 1Y', 'Decline 2Y', 'Decline 5Y', 'Decline 10Y']
                
                for col, header in enumerate(headers, 1):
                    cell = summary_ws.cell(1, col, header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write summary data
                row = 2
                for well_name, analysis_data in sorted(selected_analyses.items()):
                    summary_ws.cell(row, 1, well_name)
                    summary_ws.cell(row, 2, analysis_data.get('district', ''))
                    summary_ws.cell(row, 3, analysis_data.get('field', ''))
                    summary_ws.cell(row, 4, analysis_data.get('formation', ''))
                    
                    # Calculate decline percentages using FITTED MODEL VALUES
                    applied_wells = analysis_data.get('applied_wells', [well_name])
                    
                    # Get model parameters
                    qi = None
                    Di = None
                    b = None
                    
                    # Try to extract from results text
                    results_text = analysis_data.get('results_text', '')
                    for line in results_text.split('\n'):
                        if 'qi =' in line and 'Di =' in line and 'b =' in line:
                            try:
                                parts = line.split(',')
                                qi = float(parts[0].split('=')[1].strip())
                                Di = float(parts[1].split('=')[1].strip())
                                b = float(parts[2].split('=')[1].strip())
                            except:
                                pass
                    
                    # Calculate t values for the model
                    t_values = None
                    if len(applied_wells) > 1:
                        # For aggregated wells, use aggregated data
                        df_aggregated = self.aggregate_well_data(applied_wells)
                        if not df_aggregated.empty:
                            df_aggregated = df_aggregated.sort_values('Prod_Date')
                            df_aggregated['Prod_Date'] = pd.to_datetime(df_aggregated['Prod_Date'])
                            df_aggregated['t'] = (df_aggregated['Prod_Date'] - df_aggregated['Prod_Date'].min()).dt.days // 30
                            t_values = df_aggregated['t'].values
                    else:
                        # Single well
                        df_well = self.df_master[self.df_master['Well_Name'] == applied_wells[0]].copy()
                        if not df_well.empty:
                            df_well = df_well.sort_values('Prod_Date')
                            df_well['Prod_Date'] = pd.to_datetime(df_well['Prod_Date'])
                            df_well['t'] = (df_well['Prod_Date'] - df_well['Prod_Date'].min()).dt.days // 30
                            t_values = df_well['t'].values
                    
                    # Calculate decline percentages using model values
                    decline_pcts = []
                    if qi is not None and Di is not None and b is not None and t_values is not None:
                        # Get the total months of history data
                        applied_wells_for_data = analysis_data.get('applied_wells', [well_name])
                        
                        # Calculate total history months
                        if len(applied_wells_for_data) > 1:
                            df_for_months = self.aggregate_well_data(applied_wells_for_data)
                        else:
                            df_for_months = self.df_master[self.df_master['Well_Name'] == applied_wells_for_data[0]].copy()
                        
                        if not df_for_months.empty:
                            df_for_months = df_for_months.sort_values('Prod_Date')
                            df_for_months['Prod_Date'] = pd.to_datetime(df_for_months['Prod_Date'])
                            total_months = len(df_for_months)
                            total_years = total_months / 12
                        else:
                            total_months = 0
                            total_years = 0
                        
                        # Calculate declines for each period
                        for years in [1, 2, 5, 10]:
                            if total_years >= years:
                                # Enough data for this period
                                decline_str = self.calculate_decline_percentage_model(t_values, qi, Di, b, years)
                                if decline_str:
                                    # Normalize per year (divide by number of years)
                                    decline_pct = float(decline_str) / years
                                    decline_pcts.append(f"{decline_pct:.2f}")
                                else:
                                    decline_pcts.append('')
                            elif total_months >= 6 and years == 1:
                                # Less than 1 year but at least 6 months - scale to 1 year and add *
                                decline_str = self.calculate_decline_percentage_model(t_values, qi, Di, b, total_months / 12)
                                if decline_str:
                                    # Scale to 1 year and add star
                                    decline_pct = float(decline_str) / (total_months / 12)
                                    decline_pcts.append(f"{decline_pct:.2f}*")
                                else:
                                    decline_pcts.append('')
                            else:
                                # Not enough data
                                decline_pcts.append('')
                    else:
                        decline_pcts = ['', '', '', '']
                    
                    # Write decline percentages (starting at column 5)
                    for idx, decline_value in enumerate(decline_pcts, 5):
                        summary_ws.cell(row, idx, decline_value)
                    
                    row += 1
                
                # Auto-size columns for summary sheet
                for col in range(1, len(headers) + 1):
                    summary_ws.column_dimensions[get_column_letter(col)].width = 18
            
            # Save workbook
            wb.save(filepath)
            
            QMessageBox.information(self, "Success", 
                                   f"Export completed successfully!\n\n"
                                   f"File: {Path(filepath).name}\n"
                                   f"Fields: {len(field_data)}\n"
                                   f"Analyses: {len(selected_analyses)}")
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")
            import traceback
            traceback.print_exc()

    def _build_selection_title(self, well_names):
        """Build a grouping-aware title for the provided wells."""
        if not well_names:
            return ""
        if len(well_names) == 1:
            return f"Well: {well_names[0]}"
        selected_set = set(well_names)
        df_source = getattr(self, 'df_master', None)
        if df_source is None:
            df_source = self.df_all
        df_unique = df_source[[
            'Well_Name', 'District', 'Field', 'Alias_Formation'
        ]].drop_duplicates(subset=['Well_Name'])
        df_selected = df_unique[df_unique['Well_Name'].isin(selected_set)]
        # Formation level
        unique_triples = df_selected[['District', 'Field', 'Alias_Formation']].drop_duplicates()
        if len(unique_triples) == 1:
            d_val = unique_triples.iloc[0]['District']
            f_val = unique_triples.iloc[0]['Field']
            fm_val = unique_triples.iloc[0]['Alias_Formation']
            full_formation_set = set(
                df_unique[
                    (df_unique['District'] == d_val) &
                    (df_unique['Field'] == f_val) &
                    (df_unique['Alias_Formation'] == fm_val)
                ]['Well_Name']
            )
            if selected_set == full_formation_set:
                d_str = str(d_val).strip()
                f_str = str(f_val).strip()
                fm_str = str(fm_val).strip()
                return f"{d_str}-{f_str}-{fm_str} ({len(well_names)} wells)"
        # Field level
        unique_pairs = df_selected[['District', 'Field']].drop_duplicates()
        if len(unique_pairs) == 1:
            d_val = unique_pairs.iloc[0]['District']
            f_val = unique_pairs.iloc[0]['Field']
            full_field_set = set(
                df_unique[
                    (df_unique['District'] == d_val) &
                    (df_unique['Field'] == f_val)
                ]['Well_Name']
            )
            if selected_set == full_field_set:
                d_str = str(d_val).strip()
                f_str = str(f_val).strip()
                return f"{d_str}-{f_str} ({len(well_names)} wells)"
        # District level
        unique_districts = df_selected[['District']].drop_duplicates()
        if len(unique_districts) == 1:
            d_val = unique_districts.iloc[0]['District']
            full_district_set = set(
                df_unique[df_unique['District'] == d_val]['Well_Name']
            )
            if selected_set == full_district_set:
                d_str = str(d_val).strip()
                return f"{d_str} ({len(well_names)} wells)"
        # Fallback
        return f"Aggregated Production ({len(well_names)} wells)"

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DeclineCurveApp()
    window.show()
    sys.exit(app.exec())